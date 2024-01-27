from __future__ import annotations
import os, jdatetime
import time, math
import numpy as np
import warnings
warnings.filterwarnings("ignore")
import pickle, json
from urllib.request import urlopen
import requests
import statistics


from survey_stats.functions import *
from survey_stats.date import Date

class Data_Types:
    cross = 'cross'
    time = 'time'

class Data:
    """
    Data class include any type of data ('cross', 'time') as a table: column's labels are 'variable' and row's labels are 'index'.\n
    default type is 'cross', and TimeSeries is a child of Data that develop mathods on 'time' Data.
    """
    def __init__(self, type:str='cross', values:dict[dict]={}) -> None:
        if type == Data_Types.cross or type == Data_Types.time:
            self.type = type
        else:
            self.type = None
        self.values = values
    
    def __str__(self) -> str:
        return self.to_str()

    def dtype(self, variable:str|list[str]=[],
                  base_on_total_data:bool=False)->str|list:
        '''
        type of non-nan values:\n
        1. numeric: int or float except range of integers.\n
        2. categorical: str or range of integers for example: 1:10 or 0:9, or 70% of range of integers for example: 2,4,5,6.\n
        3. unknown: other types or a combination of numeric and string (categorical)\n
        base_on_total_data=True calls all value of variable, but \n
        base_on_total_data=False only decides based on the first 200 non-nan value.\n
        '''
        if not isinstance(variable, list):
            variable = [variable]
        if variable == []:
            variable = [v for v in self.variables()]
        index = self.index()
        res = {}
        for var in variable:
            if base_on_total_data:
                vals = {self.values[var][i] for i in index}
                vals.discard(math.nan)
            else:
                n = 0
                vals = set()
                for i in index:
                    if not is_nan(val:=self.values[var][i]):
                        vals.add(val)
                        n += 1
                    if n>=10:
                        break
            types = list({type(val) for val in vals})
            if len(types)==1:
                if types[0] == str:
                    res[var] = 'categorical'
                elif types[0] == int:
                    # min_vals, max_vals = min(vals), max(vals)
                    # if len(vals)>=len(range(min_vals, max_vals+1))*.7:
                    #     res[var] = 'categorical'
                    #     # if sorted(vals)==list(range(min_vals, max_vals+1)):
                    #     #     res[var] = 'categorical'
                    # else:
                    res[var] = 'numeric'
                elif types[0] == float:
                    res[var] = 'numeric'
                else:
                    res[var] = 'unknown'
            elif len(types)==2 and float in types and int in types:
                res[var] = 'numeric'
            else:
                res[var] = 'unknown'
        if len(variable)==1:
            if res == {}:
                raise ValueError(f"Error! there aren't any data for {variable[0]}.")
            return res[variable[0]]
        if res == {}:
            raise ValueError(f"Error! there aren't any data for {variable[0]}...{variable[-1]}.")
        return res

    def set_dtype(self, var_type:dict):
        '''
        You can change 'unknown' or any types to either 'numeric' or 'categorical' types.\n
        If you change it to a 'number', non-numerics (string, date, etc.) become 'nan'.
        If you change to a 'categorical', the non-strings are converted to 'str' with the str() function.
        '''
        index = self.index()
        for var in var_type:
            if var_type[var] == 'numeric':
                for i in index:
                    if not is_nan(val:=self.values[var][i]):
                        if not (isinstance(val, float) or isinstance(val, int)):
                            self.values[var][i] = math.nan
            elif var_type[var] == 'categorical':
                for i in index:
                    if not is_nan(val:=self.values[var][i]):
                        self.values[var][i] = str(val)

    def to_str(self, variable:list|str=[], full_variables:bool=False,
                     index:list=[], full_index:bool=False,
                     decimals:int=4, max_char:int=20, formated:bool=True)->str:
        '''
        convert to a simple table as a string.

        1. variables: if it's empty all variables will be selected.
        2. full_variables: True if we want to display all variables.
        3. index: if it's not empty, only this index will be selected.
        4. full_index: True if we want to display all index.
        5. decimals: count of decimals for numeric values.
        6. formated: True if we want use ',' in numeric values.
        '''
        if self.values == {}:
            return '[no values]'
        variables = [variable] if isinstance(variable, str) else variable
        #region vars
        if variables!=[]:
            vars = ['index'] + variables
        elif full_variables or len(self.variables())<7:
            vars = ['index'] + self.variables()
        else: 
            vars = ['index'] + self.variables()[:3] + ['...'] + self.variables()[-2:]
        #endregion
        #region inds
        if index!=[]:
            inds = index
        elif full_index or len(self.index())<15:
            inds = self.index()
        else:
            inds = self.index()[:10] + ['⁝'] + self.index()[-5:]
        #endregion
        #region widths[var]
        widths = {}
        for var in vars:
            for i in inds:
                if i!='⁝' and var!='...' and var!='index':
                    if var in self.variables() and i in self.index():
                        v = self.values[var][i]
                        if formated:
                            v = to_formated_str(v,decimals)[:max_char]
                        else:
                            v = str(v)[:max_char]
                        if not var in widths.keys():
                            widths[var] = max(len(v), len(str(var)))
                        else:
                            widths[var] = max(widths[var], len(v))
                    elif not var in self.variables():
                        raise ValueError(f"Error! '{var}' is not in variables.")
                    elif not i in self.index():
                        raise ValueError(f"Error! '{i}' is not in index.")
                elif var=='index':
                    if not var in widths.keys():
                        widths[var] = len(str(i))
                    else:
                        widths[var] = max(widths[var], len(str(i)))
                elif var=='...':
                    if not var in widths.keys():
                        widths[var] = 3
                    else:
                        widths[var] = max(widths[var], 3)
            widths[var] += 4
        #endregion
        rows = f'type: {self.type}\n'
        total_width = sum([widths[var] for var in vars])
        #region title
        rows += '\n ' + '-' * (total_width + len(vars)-1) + '\n|'
        for var in vars:
            rows += str(var).center(widths[var]) + '|'
        #endregion
        rows += '\n ' + '-' * (total_width + len(vars)-1)
        #region values
        for i in inds:
            for var in vars:
                if var == 'index':
                    rows += '\n|' + str(i).center(widths[var]) + '|'
                elif var == '...':
                    rows += '...'.center(widths['...']) + '|'
                elif i in self.index():
                    v = self.values[var][i]
                    if formated:
                        v = to_formated_str(v,decimals)[:max_char]
                    else:
                        v = str(v)[:max_char]
                    rows += v.center(widths[var]) + '|'
                elif i == '⁝':
                    rows += '⁝'.center(widths[var]) + '|'
        #endregion
        rows += '\n ' + '-' * (total_width + len(vars)-1)
        rows += f'\n{len(self.index())} × {len(self.variables())}'
        return rows

    def variables(self) -> list[str]:
        '''
        return list of variables.
        '''
        return list(self.values)

    def items(self, deep:bool=False) ->list[tuple]:
        '''return values as a list of tuple'''
        if deep:
            return [(v, [(i,x) for i,x in vals.items()]) for v, vals in self.values.items()]
        return [(v, vals) for v, vals in self.values.items()]

    def index(self, without_checking:bool=True) -> None:
        '''
        return list of index.
        if without_checking == True, then return index of first variable, otherwise will be checked index of all variables equal to index of first variable.
        '''
        try:
            index = list(self.values[self.variables()[0]])
            if not without_checking:
                for v in self.variables()[1:]:
                    if index != list(self.values[v]):
                        raise ValueError("Error! index aren't same for all variables.")
            return index
        except:
            self.fix_index()
            try:
                return self.index()
            except:
                return []

    def fix_index(self)->None:
        index = set()
        for var in self.variables():
            index.update(self.values[var].keys())
        for var in self.variables():
            for i in index.difference(self.values[var].keys()):
                self.values[var][i] = math.nan
                
    def set_index(self, variable:str|list, drop_variable:bool=True, index_name:str='',
                  print_progress:bool=False, indent:int=0) -> None:
        '''
        setting values of a variable in data or a list out of data as index of all variables.\n
        warning: If the index of a variable has duplicate values, the duplicates will be removed.\n
        example:
        a = Data(values={'v1':{'1':'ali', '2':'ali', '3':'hosein'},\n
                 'v2':{'1':128, '2':153, '3':190}})\n
        a = Data(values={'v1':{'1':'ali', '2':'ali', '3':'hosein'}})\n
        print(a)\n
        a.set_index('v1',drop_var=False)\n
        print(a)\n
        '''
        if print_progress:
            print(' '*indent+'setting index: ', end='')
        index = self.index()
        if not isinstance(variable, list):
            if print_progress:
                print(variable)
            if variable in self.variables():
                self.values = {v:{self.values[variable][i]:self.values[v][i]
                                  for i in self.values[variable].keys()} 
                                    for v in self.variables()
                                        if not (v == variable and drop_variable)}
            else:
                raise ValueError(f"Error! {variable} is not in variables of data.")
        elif isinstance(variable, list):
            if print_progress:
                print(variable[0],'...',variable[-1])
            if len(variable) == len(self.index()):
                self.values = {v:{variable[i]:val
                                  for i, val in enumerate(self.values[v].values())}
                                    for v in self.variables()}
            else:
                raise ValueError(f"Error! len of var (={len(variable)}) is not equal to len of data index (={len(self.index())}).")
        else:
            return
        if index_name != '':
            for v in self.values:
                new_index = self.values[v].keys()
            self.values.update({index_name:{i_new:index[i] for i,i_new in enumerate(new_index)}})

    def set_names(self, old_name:str|list[str]=[], new_name:str|list[str]=[],
                  print_progress:bool=False, indent:int=0)->None:
        '''
        rename variables.
        '''
        if print_progress:
            print(' '*indent + 'setting names:',end='')
        if not(isinstance(new_name, list) and isinstance(old_name, list)):
            if print_progress:
                print(old_name+'->'+new_name)
            self.values.update({new_name:self.values.pop(old_name)})
        elif isinstance(new_name, list) and isinstance(old_name, list) and len(new_name)==len(old_name):
            if print_progress:
                print(f'[{old_name[0]}...]->[{new_name[0]}...]')
            self.values.update({new_name[i]:self.values.pop(v) for i, v in enumerate(old_name)})
        else:
            raise ValueError(f"Error! size of old_name ({len(old_name)}) and new_name ({len(new_name)}) is not equal.")

    def select_variables(self, variable:str|list[str]=[],
                  print_progress:bool=False, indent:int=0) -> Data:
        '''
        return values of selected variables as a Data
        '''
        if print_progress:
            print(' '*indent +'selecting variable:', end='')
        if not isinstance(variable, list):
            if print_progress:
                print(variable)
            return Data(self.type,{variable:self.values[variable].copy()})
        elif isinstance(variable, list):
            if variable!=[]:
                if print_progress:
                    print(variable[0],'...',variable[-1])
                return Data(self.type,{v:self.values[v].copy() for v in variable})
        else:
            raise ValueError(f"Error! [{vars[:1]},...,{vars[-1:]}] is not a list of variables.")

    def select_index(self,index:list|str|int|float,
                  print_progress:bool=False, indent:int=0)->Data:
        '''return values of selected index as a Data'''
        if print_progress:
            print(' '*indent+ 'selecting index:', end='')
        if not isinstance(index, list):
            if print_progress:
                print(index)
            if index in self.index():
                res_dict = {}
                for var in self.variables():
                    if not var in res_dict:
                        res_dict[var] = {}
                    res_dict[var][index] = self.values[var][index]
                index.remove(i)
                return Data(type=self.type, values=res_dict)
            raise ValueError(f"Error! {index} is not in index of data.")
        elif isinstance(index, list):
            if print_progress:
                print(index[0]+'...'+index[-1])
            res_dict = {}
            index_copy = index.copy()
            for i in self.index():
                if i in index_copy:
                    for var in self.variables():
                        if not var in res_dict:
                            res_dict[var] = {}
                        res_dict[var][i] = self.values[var][i]
                    index_copy.remove(i)
            return Data(type=self.type, values=res_dict)
        else:
            raise ValueError(f"Error! {index} is not a list of index.")

    def drop(self, variable_name:str|list[str], print_error:bool=True,
                  print_progress:bool=False, indent:int=0)->None:
        '''remove one or more variables from values'''
        if print_progress:
            print(' '*indent+'drop variable:', end='')
        if not isinstance(variable_name, list):
            if print_progress:
                print(variable_name)
            self.values.pop(variable_name)
        elif isinstance(variable_name, list):
            if print_progress:
                print(variable_name[0]+'...'+variable_name[-1])
            no_vars = []
            for var in variable_name:
                if var in self.variables():
                    self.values.pop(var)
                else:
                    no_vars.append(var)
            if print_error and no_vars != []:
                print(' '*indent+f"Error! there arn't {no_vars} in data.")
        else:
            raise ValueError(f"Error! '{variable_name}' is not a list of variables.")

    def drop_index(self, index:list|any,
                  print_progress:bool=False, indent:int=0, subindent:int=5)->None:
        '''remove one or more index from values'''
        if print_progress:
            print(' '*indent+'dropping index:', end='')
            if not isinstance(index, list):
                print(index)
            else:
                if len(index)==1:
                    print(index[0])
                elif len(index)>1:
                    print(index[0]+'...'+index[-1])
            start, prelog, n = time.perf_counter(), 0, len(self.variables())
        if not isinstance(index, list):
            index = [index]
        index_copy = index.copy()
        for j, var in enumerate(self.variables()):
            for i in index_copy:
                self.values[var].pop(i)
            if print_progress:
                prelog = progress(start, j, n, prelog, var, indent=indent+subindent)
    
    def add_a_dummy(self, condition:list[list[tuple]], name:str='', add_to_data:bool=False,
                  print_progress:bool=False, indent:int=0)->Data:
        '''
        create a variable by use of variables on data and then add to data (if add_to_data==True) or only return\n
        condition = [('sex','=','female'), ('age','<',20)]\n
        names = sex=female_and_age<20\n
        '''
        if print_progress:
            if len(condition)>1:
                print(' '*indent+'adding a dummy:',condition[0],'...')
            else:
                print(' '*indent+'adding a dummy:',condition[0])
        dummy_values = {}
        for i in self.index():
            satisfied, isnan = True, False
            for var, sign, val in condition:
                if not is_nan(self.values[var][i]):
                    satisfied = satisfied and check_condition(self.values[var][i], sign, val)
                else:
                    isnan = True
                    break
            if satisfied:
                dummy_values[i] = 1
            elif not isnan:
                dummy_values[i] = 0
            else:
                dummy_values[i] = math.nan
        if name == '':
            name = condition[0][0] + condition[0][1] + str(condition[0][2])
            for var, sign, val in condition[1:]:
                name += '_and_' + var + sign + str(val)
        res = {name:dummy_values}
        if add_to_data:
            self.values.update(res)
        return Data(self.type, res)
            
    def add_dummies(self, conditions:list[dict], names:list=[], add_to_data:bool=False,
                  print_progress:bool=False, indent:int=0)->Data:
        '''
        create some variables by use of variables on data and then add to data (if add_to_data==True) or only return\n
        conditions = [[('sex','=','female'),('age','<',20)],[()],[()],...,[()]]\n
                               |_________a condition__________| |__| |__|     |__|\n
        names =                      sex=female_age<20     ...
        '''
        if print_progress:
            print(' '*indent+'adding dummies:', conditions[0][0],'...')
        if len(names) != len(conditions):
            names = [''] * len(conditions)
        values = {}
        for cond in conditions:
            values.update(self.add_a_dummy(cond, add_to_data).values)
        return Data(self.type, values)

    def dropna(self, variable:list[str]|str=[], print_error:bool=True,
                  print_progress:bool=False, indent:int=0, subindent:int=5)->None:
        '''drop index that minimum value of a variable is nan.'''
        if print_progress:
            print(' '*indent+'dropping nans')
            start, prelog = time.perf_counter(), 0
        if isinstance(variable, str):
            variable = [variable]
        no_variables, variables = [], []
        for v in variable:
            if v in self.variables():
                variables.append(v)
            else:
                no_variables.append(v)
        if no_variables != [] and print_error:
            print(' '*indent + f"Error! {no_variables} is not in data variables.")
        variables = variables if variables != [] else self.variables()
        n = len(self.index())
        for j, i in enumerate(self.index()):
            for v in variables:
                if is_nan(self.values[v][i]):
                    for vi in self.variables():
                        self.values[vi].pop(i)
                    break
            if print_progress:
                prelog = progress(start, j, n, prelog, i,True, indent+subindent)
    
    def drop_all_na(self, variable:list[str]|str=[], print_error:bool=True,
                  print_progress:bool=False, indent:int=0, subindent:int=5)->None:
        '''drop index that value of all variable or all value of a variable is nan.'''
        if print_progress:
            print(' '*indent+'dropping nans')
            start, prelog = time.perf_counter(), 0
        if isinstance(variable, str):
            variable = [variable]
        no_variables, variables = [], []
        for v in variable:
            if v in self.variables():
                variables.append(v)
            else:
                no_variables.append(v)
        if no_variables != [] and print_error:
            print(' '*indent + f"Error! {no_variables} is not in data variables.")
        variables = variables if variables != [] else self.variables()
        n = len(self.index())
        if print_progress:
            print(' '*(indent+subindent)+'dropping rows that all variables are nans')
        for j, i in enumerate(self.index()):
            nan = True
            for v in variables:
                nan = nan and is_nan(self.values[v][i])
            if nan:
                for vi in self.variables():
                    self.values[vi].pop(i)
            if print_progress:
                prelog = progress(start, j, n, prelog, i,True, indent+2*subindent)
        n, start, prelog = len(variable), time.perf_counter(), 0
        if print_progress:
            print(' '*(indent+subindent)+'dropping variables that all their values are nans')
        for j, v in enumerate(variables):
            nan = True
            for i in self.index():
                nan = nan and is_nan(self.values[v][i])
            if nan:
                self.drop(v)
            if print_progress:
                prelog = progress(start, j, n, prelog, i, True, indent+2*subindent)

    def value_to_nan(self, value:str|int|float, variable:list[str]|str=[],
                     print_error:bool=True,
                  print_progress:bool=False, indent:int=0, subindent:int=5)->None:
        '''
        Puts nan in place of a value like '-' or ''.
        '''
        if print_progress:
            print(' '*indent+f"convert '{value}'s to nan")
            start, prelog = time.perf_counter(), 0
        if isinstance(variable, str):
            variable = [variable]
        no_variables, variables = [], []
        for v in variable:
            if v in self.variables():
                variables.append(v)
            else:
                no_variables.append(v)
        if no_variables != [] and print_error:
            print(' '*indent + f"Error! {no_variables} is not in data variables.")
        if variables == []:
            variables = self.variables()
        n = len(variables)
        for j,var in enumerate(variables):
            for i in self.index():
                if self.values[var][i] == value:
                    self.values[var][i] = math.nan
            if print_progress:
                prelog = progress(start,j,n,prelog,var, True,indent+subindent)

    def to_numpy(self, vars:list[str]=[],
                  print_progress:bool=False, indent:int=0)->None:
        '''
        convert values to numpy array.
        '''
        if print_progress:
            print(' '*indent+"values to numpy array.")
        lst = []
        for i in self.index():
            in_lst = []
            for var in self.values.keys():
                if (var in vars) or (vars==[]):
                    if is_numeric(self.values[var][i]):
                        in_lst.append(self.values[var][i])
                    else:
                        raise ValueError(f"Error! value of {self.values[var][i]} is not numeric.")
            lst.append(in_lst)
        return np.array(lst)

    def add_data(self, new_data:Data=None,
                  print_progress:bool=False, indent:int=0, subindent:int=5)->None:
        '''
        update values by new data.
        '''
        if print_progress:
            print(' '*indent+"adding data.")
            start, prelog = time.perf_counter(), 0
        old_index = self.index().copy()
        if self.index() == []:
            new_index = new_data.index()
            indexes = new_index
        else:
            if new_data.values != {}:
                new_index = new_data.index().copy()
            else:
                new_index = []
            indexes = list(set(old_index+new_index))
            try:
                indexes.sort()
            except:
                pass
        new_index = list(set(new_index)-set(old_index))
        old_vars = self.variables()
        new_vars = new_data.variables()
        vars = old_vars + [var for var in new_vars if not var in old_vars]
        n = len(vars)
        for j,var in enumerate(vars):
            if not var in old_vars:
                self.values[var] = dict(zip(indexes,[math.nan]*len(indexes)))
            if var in new_vars:
                new_values = new_data.values[var].copy()
                for i, v in new_data.values[var].items():
                    try:
                        if is_nan(v) and self.values[var][i]:
                            new_values.pop(i)
                    except:
                        pass
                self.values[var].update(new_values)
            elif new_index != []:
                new_vals = dict(zip(new_index,[math.nan]*len(new_index)))
                self.values[var].update(new_vals)
            if print_progress:
                prelog = progress(start, j, n, prelog, var, True, indent+subindent)
    
    def transpose(self,
                  print_progress:bool=False, indent:int=0, subindent:int=5)->Data:
        '''
        convert index to variables and variables to index.
        '''
        if print_progress:
            print(' '*indent+'transpose data')
            start, prelog, n,j= time.perf_counter(),0,len(self.variables()),0
        values_t = {}
        for var, ival in self.values.items():
            for i, val in ival.items():
                if i in values_t.keys():
                    values_t[i][var] = val
                else:
                    values_t[i] = {var:val}
            if print_progress:
                prelog = progress(start, j,n,prelog,var, True,indent+subindent)
                j+=1
        return Data(self.type, values_t)

    def __len__(self):
        '''the count of index that all variables are not nan.'''
        return len([i for i in self.index()
                   if len([v for v in self.variables() if is_nan(self.values[v][i])])==0])
    
    def count(self, variables:str|list=[], index:list=[], reverse:bool=False)->int:
        '''
        reverse=False -> the count of index that values of all variables are not nan.\n
        reverse=True -> the count of variables that values of them in all index are not nan.\n
        '''
        if variables==[]:
            variables = self.variables()
        if isinstance(variables,str):
            variables = [variables]
        if index==[]:
            index = self.index()
        if reverse:
            n = 0
            for v in variables:
                for i in index:
                    if is_nan(self.values[v][i]):
                        break
                else:
                    n += 1
            return n
        n = 0
        for i in index:
            for v in variables:
                if is_nan(self.values[v][i]):
                    break
            else:
                n += 1
        return n

    def add_trend(self, print_progress:bool=False, indent:int=0):
        '''
        add a variable as trend to data.
        '''
        if print_progress:
            print(' '*indent + 'adding trend')
        j = 0
        self.values['trend'] = {i:j+1 for j, i in enumerate(self.index())}

    fillna_methods = ['last', 'mean', 'growth']
    def fillna(self, value:str|int|float='', variable:str|list[str]=[] , replace:bool=False,
                  print_progress:bool=False, indent:int=0):
        '''
        Put a value like '-' or '' on place of nans. if replace==False create a new variable.
        '''
        if print_progress:
            print(' '*indent+'filling NaNs.', end='')
        if type(variable) == str:
            if print_progress:
                print(variable)
            if not replace:
                self.values[variable+'_1'] = {}
            for i in self.index():
                if replace:
                    if is_nan(self.values[variable][i]):
                        self.values[variable][i] = value
                else:
                    if is_nan(self.values[variable][i]):
                        self.values[variable+'_1'][i] = value
                    else:
                        self.values[variable+'_1'][i] = self.values[variable][i]
        elif type(variable)==list:
            if variable==[]:
                variable = self.variables()
            if print_progress:
                print(variable[0],'...',variable[-1])
            if variable == []:
                variable = self.variables()
            for var in variable:
                self.fillna(value, var, replace)

    def fill(self, variable:str|list[str]=[], value_find:any='', value_replace:any='',
                  print_progress:bool=False, indent:int=0)->Data:
        '''
        put value_replace on places of value_find.
        '''
        if print_progress:
            print(' '*indent+f"replacing '{value_replace}' on '{value_find}' in ", end='')
            if not isinstance(variable, list):
                print(variable)
            else:
                print(variable[0],'...',variable[-1])
        variables = [variable] if isinstance(variable, str) else variable
        if variables == []:
            variables = self.variables()
        for i in self.index():
            for var in variables:
                if self.values[var][i] == value_find:
                    self.values[var][i] = value_replace

    def sort(self, key:str='', ascending:bool=True,
                  print_progress:bool=False, indent:int=0):
        '''
        sort values by index or value of a variable.\n
        noice: this method only affect on print.
        '''
        if key == '':
            key = 'index'
        if print_progress:
            method = 'ascending' if ascending else 'decending'
            print(' '*indent+f"sorting data by {method} on {key}.")
        if key=='index':
            for var in self.variables():
                self.values[var] = dict(sorted(self.values[var].items(), reverse=not ascending))
        elif key in self.variables():
            sorted_index = [i for i, v in sorted([(i,v) for i,v in self.values[key].items()
                                                            if not is_nan(v)],
                                                            key=lambda x: x[1], reverse=not ascending)]
            sorted_index += [i for i in self.index() if is_nan(self.values[key][i])]
            for var in self.variables():
                self.values[var] = {i:self.values[var][i] for i in sorted_index}
        else:
            raise ValueError(f"Error! key ('{key}') is not in variables.")

    def add_a_variable(self, name:str, values:list,
                  print_progress:bool=False, indent:int=0):
        '''
        add a variable by values only to data.
        '''
        if print_progress:
            print(' '*indent+f"adding '{name}' by {self.variables()[0]}={values[0]}...{self.variables()[-1]}={values[-1]}")
        if len(self.index())==len(values):
            self.values[name] = dict(zip(self.index(), values))
        else:
            raise ValueError(f"Error! lenght of values ({len(values)}) must be equal to lenght of index ({len(self.index())}).")

    def to_timeseries(self, print_progress:bool=False, indent:int=0, subindent:int=5):
        '''
        convert data to a timeseries.
        '''
        if print_progress:
            print(' '*indent+'converting type from Data to TimeSeries')
        data = TimeSeries('time', self.values)
        data.complete_dates(print_progress=print_progress, indent=indent+subindent, subindent=subindent)
        return data

    def line_plot(self, vars:list[str], title:str='', show:bool=True, print_progress:bool=False, indent:int=0):
        '''
        A graph plots one or more variables.\n
        matplotlib must installed.\n
        pip install matplotlib.
        '''
        if print_progress:
            print(' '*indent+"ploting variables: "+', '.join(vars))
        from matplotlib import pyplot as plt
        fig, ax = plt.subplots(1,1)
        plt.xticks(rotation='vertical')
        plt.title(title)
        plt.margins(0.1)
        plt.subplots_adjust(left=0.1, right=0.9, top=0.9, bottom=0.2)
        legs = []
        for var in vars:
            ax.plot(self.index(), list(self.values[var].values()))
            legs.append(var)
        ax.legend(legs)
        interval = int(len(self)/20) if len(self)>40 else 2 if len(self)>40 else 1
        ax.set_xticks(ax.get_xticks()[::interval])
        plt.tight_layout()
        if show:
            plt.show()
        return plt

    def add_index(self, index:list, print_progress:bool=False, indent:int=0):
        '''
        add index to end of index and put nan for values of variables.
        '''
        if print_progress:
            if len(index)>1:
                print(' '*indent+f"adding index to data: {index[0]}...{index[-1]}")
            else:
                print(' '*indent+f"adding index to data: {index[0]}")
        for var in self.variables():
            for i in index:
                if not i in self.values[var].keys():
                    self.values[var][i] = math.nan

    def add_a_group_variable(self, variable:str, map:list|dict, other:any='', name:str='')->None:
        '''
        if 'variable' is 'numeric', then 'map' is a list of thresholds.\n
        for example:\n
        - 'map' is [0.5] -> we have 2 groups: x<=0.5 and x>0.5\n
        - 'map' is [0.5, 1.5] -> we have 3 groups: x<=0.5 and 0.5<x<=1.5 and 1.5<x.\n
        \n
        if 'variable' is 'categorical', then 'map' is a dictionary of mapping.\n
        for example: 'map' is\n
        - {'male':0,'female':1}
        - {'illiterate':'non-academic', 'primary':'non-academic',
                'diploma':'non-academic', 'associate':'academic',
                'bachelor':'academic', 'ma':'academic', 'PhD':'academic'}
        - {'strongly disagree':-1, 'disagree':-.5, 'neutral':0, 'agree':.5, 'strongly agree':1}\n
        Note: if map incomplete, for example for education degress we have \n
        map = {'illiterate':'non-academic', 'primary':'non-academic', 'diploma':'non-academic'}\n
        if other=='', variable values put on it, but other != '' for example 'academic' this value put on it.\n
        '''
        if name=='':
            if not f'{variable}_g' in self.variables():
                name = f'{variable}_g'
            else:
                k = 1
                while f'{variable}_g{k}' in self.variables():
                    k += 1
                name = f'{variable}_g{k}'
        elif name in self.variables():
            k = 1
            while f'{name}_g{k}' in self.variables():
                k += 1
            name = f'{name}_g{k}'
        if self.dtype(variable) == 'numeric':
            if isinstance(map, list):
                for x in map:
                    if not is_numeric(x):
                        raise ValueError(f"error! '{x}' in map is not numeric.")
                map.sort()
                self.values[name] = {}
                for i in self.index():
                    if is_nan(vali:=self.values[variable][i]):
                        self.values[name][i] = math.nan
                    else:
                        self.values[name][i] = sum([val<vali for val in map if not is_nan(val)])
            else:
                raise ValueError(f"error! '{map}' is not list.")
        else:
            if isinstance(map, dict):
                vals = set(self.values[variable].values())
                if math.nan in vals:
                    vals.remove(math.nan)
                for v in map:
                    if not v in vals:
                        raise ValueError(f"Error! '{v}' in keys of map is not in values of {variable}.")
                self.values[name] = {}
                for i in self.index():
                    for v in map:
                        if self.values[variable][i] == v:
                            self.values[name][i] = map[v]
                            break
                    else:
                        self.values[name][i] = other if other!='' else self.values[variable][i]
            else:
                raise ValueError(f"error! '{map}' is not dict.")

    def to_float(self, vars:list[str]=[]):
        if vars==[]:
            vars = self.variables()
        for var in vars:
            for i in self.index():
                g = self.values[var][i]
                try:
                    self.values[var][i] = float(g)
                except:
                    self.values[var][i] = math.nan

    #region reading and writing
    #region pickle
    @classmethod
    def load(cls, path_file:str)->Data:
        with open(path_file, mode='rb') as f:
            return pickle.load(f)

    def dump(self, path_file:str)->None:
        with open(path_file, mode='wb') as f:
            pickle.dump(self, f)
    #endregion
    #region text
    @classmethod
    def read_text(cls, path_file:str, data_type:str='cross', na:any='', index:str='index', 
                  variable_number_range:tuple[int]=(),
                  variable_names:list[str] = [], only_names:bool=False,
                  seprator:str=',', encoding:str='utf8',
                  print_progress:bool=False, indent:int=0, subindent:int=5)->list|Data:
        '''
        read data from a txt file.\n
        if there is a larg file more than RAM, suggest reading a part of file.\n
        two way for reading a part of file:\n
        1. range of number of variable (1,count of variables)\n
        2. variable names
        '''
        if print_progress:
            print(' '*indent+"reading file: ", end='')
            start_time, prelog = time.perf_counter(), 0
        data = cls(data_type, values={})
        if path_file[:4].lower()=='http':
            if print_progress:
                print(path_file[:8]+path_file[8:].split('/')[0])
            response = urlopen(path_file)
            lines = [l.decode(encoding) for l in response.readlines()]
            n = len(lines)
            vars, is_first = [], True
            for i, line in enumerate(lines):
                if is_first:
                    vars = [v.strip().replace('ï»؟','').replace('\n','').replace('\ufeff','').replace('?','') for v in line.split(seprator)]
                    for j,var in enumerate(vars):
                        if var == index:
                            data.values[(index:=var+'_')] = {}
                            vars[j] = index
                        else:
                            data.values[var] = {}
                    is_first = False
                else:
                    vals = [v.strip().replace('ï»؟','').replace('\n','').replace('\ufeff','').replace('?','') for v in line.split(seprator)]
                    for j, val in enumerate(vals):
                        if val.strip() == na:
                            data.values[vars[j]][i] = math.nan
                        elif val.isdigit():
                            data.values[vars[j]][i] = int(val)
                        elif is_numeric_str(val):
                            data.values[vars[j]][i] = to_float(val)
                        else:
                            data.values[vars[j]][i] = val.strip()
                if print_progress:
                    prelog = progress(start_time, i, n, prelog, indent=indent+subindent)
            if index != 'index':
                data.set_index(index)
        else:
            if print_progress:
                print(path_file.split('\\')[-1])
            with open(path_file,'r', encoding=encoding) as f:
                lines = f.readlines()
                n = len(lines)
                start = True
                for i, line in enumerate(lines):
                    if start:   # variable names in first line
                        vars = [v.strip().replace('0xd4','').replace('ï»؟','').replace('\n','').replace('\ufeff','').replace('?','') for v in line.split(seprator)] # include index if index in vars
                        vars_names = [v for v in vars if v != index]
                        if only_names:
                            return vars_names
                        index_i = vars.index(index) if index in vars else -1
                        if variable_number_range != () and len(variable_number_range)==2:
                            min_no, max_no = variable_number_range
                            if type(min_no) != int and type(max_no)==int:
                                err = f'variable_number_range {variable_number_range} must be a tuple of two int.'
                                raise ValueError(err)
                        if variable_names != []:
                            variable_names_bool = dict([(var,True) if var in variable_names else (var,False) for var in vars])
                        start = False
                    else:       # values in other lines
                        vals = [val.replace('ï»؟','').replace('\n','').replace('\ufeff','').replace('?','').strip() for val in line.split(seprator)]
                        index_val = vals[vars.index(index)] if index in vars else str(i+1)
                        if index_val.isdigit():
                            index_val_ = int(index_val)
                        elif is_numeric_str(index_val):
                            index_val_ = to_float(index_val)
                        else:
                            index_val_ = index_val.strip()
                        for j, val in enumerate(vals):
                            if len(vals)!=len(vars):
                                raise ValueError(f"Error in {i+1}th line. length of line ({len(vals)}) is not equal to length of variable names ({len(vars)}). line text is:\n'{line[:-1]}'\nline spilits are:\n{vals}\nvariable names are:\n{vars}")
                            if variable_number_range != () and len(variable_number_range)==2:
                                cond = (min_no <= j <= max_no) and j != index_i
                            elif variable_names != []:
                                cond = variable_names_bool[vars[j]]
                            else:
                                cond = (j != index_i) or (not index in vars)
                            if cond:
                                val = val.replace('ï»؟','').replace('\n','')
                                try:
                                    val = val.strip()
                                except:
                                    pass
                                try:
                                    if not vars[j] in data.values:
                                        data.values [vars[j]] = {}
                                except:
                                    raise ValueError(f'Error in reading file "{path_file[:8]}". in line {j+1}: splited line ({vals}) is not equal to first line ({vars})')
                                if val == na:
                                    data.values[vars[j]][index_val_] = math.nan
                                elif val.isdigit():
                                    data.values[vars[j]][index_val_] = int(val)
                                elif is_numeric_str(val):
                                    data.values[vars[j]][index_val_] = to_float(val)
                                else:
                                    data.values[vars[j]][index_val_] = val
                    if print_progress:
                        prelog = progress(start_time, i, n, prelog, indent=indent+subindent)
        return data

    def to_text(self, path_file:str, na:str='', replace:bool = True, skip_index:bool=False, seprator:str=',',
                  print_progress:bool=False, indent:int=0, subindent:int=5):
        '''
        save data to a txt file.
        '''
        if print_progress:
            file_name = path_file.split('\\')[-1]
            print(' '*indent+f'saving data to {file_name}.')
        if os.path.exists(path_file):
            if replace:
                os.remove(path_file)
            else:
                old_name = path_file.split('\\')[-1]
                new_name = input(f"there is a file with same name '{old_name}', please, enter a new name without the path: ")+'.csv'
                path_file = path_file.replace(old_name,new_name)

        with open(path_file, 'a', encoding='utf-8') as f:
            if skip_index:
                title = ''
            else:
                title = 'index'
            start = True
            for var in self.variables():
                if not(start and skip_index):
                    title += seprator
                start = False
                title += str(var.replace(seprator, ''))
            f.write(title + '\n')
            if print_progress:
                start_time, prelog, n = time.perf_counter(), 0, len(self.index())
            for j, i in enumerate(self.index()):
                if skip_index:
                    line = ''
                else:
                    line = str(i)
                start = True
                for var in self.variables():
                    if not(start and skip_index):
                        line += seprator
                    start = False
                    is_nan = False
                    if is_numeric(self.values[var][i]):
                        is_nan = math.isnan(self.values[var][i])
                    if is_nan:
                        line += na
                    else:
                        line += str(self.values[var][i]).replace(seprator, '')
                f.write(line + '\n')
                if print_progress:
                    prelog = progress(start_time, j, n, prelog,i, True,indent+subindent)

    def add_to_text(self, path_file:str,only_new_index:bool=False, na:str='', replace:bool = True, skip_index:bool=False, seprator:str=',',
                  print_progress:bool=False, indent:int=0, subindent:int=5):
        '''
        append data to end of a txt file.
        '''
        if not os.path.exists(path_file):
            self.to_csv(path_file, na, replace, skip_index)
        else:
            if print_progress:
                file_name = path_file.split('\\')[-1]
                print(' '*indent+f"adding data to {file_name}.")
                
            destination = path_file if only_new_index else path_file[:-4]+'~temp'
            mode = 'a' if only_new_index else 'w'
            with open(destination, mode, encoding='utf-8') as new_file:
                # old index
                with open(path_file,'r', encoding='utf-8') as old_file:
                    lines=old_file.readlines()
                    if print_progress:
                        if only_new_index:
                            print(' '*(indent+subindent)+f"read old index")
                        else:
                            print(' '*(indent+subindent)+f"update old index")
                        start_time, prelog, n = time.perf_counter(), 0, len(lines)
                    start = True
                    if skip_index:
                        index = 0
                    for i, line in enumerate(lines):
                        if start:   # variable names in first line
                            old_vars = [v.replace('ï»؟','').replace('\n','').replace('?','').replace('\ufeff','').replace(seprator,'') for v in line.split(seprator)]
                            new_vars = old_vars + [v.replace('\ufeff','').replace('ï»؟','').replace('\n','').replace('?','').replace(seprator,'') for v in self.variables()
                                                   if not v.replace('\ufeff','').replace('ï»؟','').replace('\n','').replace('?','').replace(seprator,'') in old_vars]
                            if not only_new_index:
                                new_file.write(seprator.join(new_vars)+'\n')
                            start = False
                        else:       # values in other lines
                            old_vals = [v.replace('ï»؟','').replace('\n','').replace(seprator,'') for v in line.split(seprator)]
                            if not skip_index:
                                index = old_vals[0]
                            else:
                                index += 1
                            if not only_new_index:
                                if index in self.index():
                                    for v in self.variables():
                                        if v in old_vars:
                                            old_vals[old_vars.index(v)] = str(self.values[v][index])
                                new_vals = old_vals + [na if math.isnan(self.values[v][index]) else str(self.values[v][index]).replace(seprator,'')
                                                       for v in self.variables() if not v in old_vars]
                                new_file.write(seprator.join(new_vals)+'\n')
                        if print_progress:
                            prelog = progress(start_time, i, n, prelog, indent=indent+2*subindent)
                # new index
                index = self.index()[self.index().index(index)+1:] if index in self.index() else self.index()
                if print_progress:
                    print(' '*(indent+subindent)+f"adding new index")
                    start_time, prelog, n = time.perf_counter(), 0, len(index)
                for j,i in enumerate(index):
                    new_vals = []
                    for v in new_vars[1:]:
                        if v in self.variables():
                            if math.isnan(self.values[v][i]):
                                new_vals.append(na)
                            else:
                                new_vals.append(str(self.values[v][i]).replace(seprator,''))
                        else:
                            new_vals.append(na)
                    new_file.write(seprator.join([str(i)]+new_vals)+'\n')
                    if print_progress:
                        prelog = progress(start_time, j, n, prelog, indent=indent+2*subindent)
            if not only_new_index:
                os.remove(path_file)
                os.rename(path_file[:-4]+'~temp', path_file)
    #endregion
    #region csv
    @classmethod
    def read_csv(cls, path_file:str, data_type:str='cross', na:any='', index:str='index',
                 variable_number_range:tuple[int]=(),
                 variable_names:list[str] = [], only_names:bool=False, encoding:str='utf8',
                 print_progress:bool=False, indent:int=0, subindent:int=5)->list|Data:
        '''
        read data from a csv file.\n
        if there is a larg file more than RAM, suggest reading a part of file.\n
        two way for reading a part of file:\n
        1. range of number of variable (1,count of variables)\n
        2. variable names
        '''
        return Data.read_text(path_file, data_type, na, index, variable_number_range, variable_names,
                              only_names, ',', encoding, print_progress, indent, subindent)

    def to_csv(self, path_file:str, na:str='', replace:bool = True, skip_index:bool=False,
                  print_progress:bool=False, indent:int=0, subindent:int=5):
        '''
        save data to a csv file.
        '''
        self.to_text(path_file, na, replace, skip_index,',',print_progress, indent, subindent)

    def add_to_csv(self, path_file:str,only_new_index:bool=False, na:str='', replace:bool = True, skip_index:bool=False,
                  print_progress:bool=False, indent:int=0, subindent:int=5):
        '''
        append data to end of a csv file.
        '''
        self.add_to_text(path_file, only_new_index, na, replace, skip_index, ',', print_progress, indent, subindent)
    #endregion
    #region xls
    @classmethod
    def read_xls(cls, path_file: str, data_type:str='cross', na:any='', index:str='index',
                 print_progress:bool=False, indent:int=0) -> Data:
        '''
        read data from a txt file.\n
        xls file is deference of excel files.\n
        xls files are html tags:<html><table><thead><body><tr><th>
        '''
        if 'http://' in path_file or 'https://' in path_file:
            file_name = path_file[:8]+path_file[8:].split('/')[0]
            try:
                content = requests.get(path_file).text
            except:
                import urllib3
                urllib3.disable_warnings()
                content = requests.get(path_file, allow_redirects=True, verify=False).text
        else:
            file_name = path_file.split('//')[-1]
            with open(path_file, encoding='utf8') as f:
                content = f.read()
        if print_progress:
            print(' '*indent+f"reading data from {file_name}")

        rows = xls_read(content)
        titles = rows[0]
        values, i = {}, 0
        for row in rows[1:]:
            cols = row
            i += 1
            val = {}
            for j, col in enumerate(cols):
                vali = col
                if vali.isdigit():
                    vali = int(vali)
                elif is_numeric_str(vali):
                    vali = to_float(vali)
                elif vali.strip() == na:
                    vali = math.nan
                val[titles[j]] = vali
            values[i] = val
        res = cls(data_type, values).transpose()
        if index in titles:
            res.set_index(index)
        elif 'index' in titles:
            res.set_index('index')
        return res

    def to_xls(self,  path_file:str, na:str='', replace:bool = True, skip_index:bool=False,
                 print_progress:bool=False, indent:int=0):
        '''
        read data from a txt file.\n
        xls file is deference of excel files.\n
        xls files are html tags:<html><table><thead><body><tr><th>
        '''
        if print_progress:
            file_name = path_file.split('//')[-1]
            print(' '*indent+'saving data to '+file_name)
        #heading
        data = '''
        <html>
            <head>
                <meta charset="utf-8" />
            </head>



            <table>
        '''
        #titles
        data += '''
            <thead>
                <tr>
                    <th>index</th>'''
        for var in self.variables():
            data += f'''
                    <th>{var}</th>'''
        data += '''
                </tr>
            </thead>
        '''
        #values
        data += '''
            <tbody>
        '''
        for row in self.index():
            data += '''
                <tr>
            '''
            data += f'''
                    <th>{row}</th>'''
            for var in self.variables():
                vali, isnan = self.values[var][row], False
                if is_numeric(vali):
                    isnan = math.isnan(vali)
                if isnan:
                    data += f'''
                        <th>{na}</th>'''
                else:
                    data += f'''
                        <th>{vali}</th>'''
            data += '''
                </tr>
            '''
        data += '''
            </tbody>
        '''
        #ending
        data += '''
            </table>
        </html>
        '''
        if os.path.exists(path_file):
            if replace:
                os.remove(path_file)
            else:
                raise ValueError(f"Error! '{path_file.split('/')[-1]}' exists.")
        with open(path_file, 'w', encoding='utf8') as f:
            f.write(data)
    #endregion
    #region excel
    @classmethod
    def read_excel_sheets(cls, path_file:str) -> list:
        '''
        return list of sheet names.\n
        'openpyxl' or 'xlrd' must installed.\n
        pip install openpyxl xlrd
        '''
        import openpyxl
        try:
            return openpyxl.load_workbook(path_file).sheetnames
        except:
            import xlrd
            return xlrd.open_workbook(path_file, on_demand=True).sheet_names()

    @classmethod
    def read_excel(cls, path_file:str, sheet:str|int=0, data_type:str='cross', na:any='', index:str='index',
                    first_row:int=0, first_col:int=0, data_only:bool=True, print_progress:bool=False, indent:int=0, subindent:int=5) -> Data:
        '''
        read data from a excel file.\n
        'openpyxl' or 'xlrd' must installed.\n
        pip install openpyxl xlrd
        '''
        if print_progress:
            file_name = path_file.split('//')[-1]
            print(' '*indent+f"reading data from {file_name}")
            start, prelog = time.perf_counter(), 0
        import openpyxl
        try:
            wb = openpyxl.load_workbook(path_file, data_only=data_only)
            if isinstance(sheet,str):
                ws = wb[sheet]
            elif isinstance(sheet,int):
                ws = wb[wb.sheetnames[sheet]]
            i = first_row
            

            values = {}
            rows = list(ws.iter_rows(first_row))
            var_names = [cell.value for cell in rows[0][first_col:]]
            for var in var_names:
                values[var] = {}
            for r, row in enumerate(rows[1:]):
                for c, cell in enumerate(row[first_col:]):
                    if (val:=cell.value)==None:
                        values[var_names[c]][r+1] = math.nan
                    else:
                        values[var_names[c]][r+1] = val
                if print_progress:
                    prelog = progress(start, r, len(rows)-1, prelog, indent=indent+subindent)
        except:
            import xlrd
            wb = xlrd.open_workbook (path_file)
            if isinstance(sheet,int):
                ws = wb.sheet_by_index(sheet)
            elif isinstance(sheet,str):
                ws = wb.sheet_by_name(sheet)
            else:
                raise ValueError(f"sheet='{sheet}' must be a int or str.")
            values = {}
            var_names = [ws.cell_value(first_row, col) for col in range(first_col, ws.ncols)]
            for var in var_names:
                values[var] = {}
            i = 0
            for row in range(first_row+1, ws.nrows):
                for col in range(first_col, ws.ncols):
                    values[ws.cell_value(first_row, col)][i]=ws.cell_value(row, col)
                if print_progress:
                    prelog = progress(start, i, ws.nrows-first_row-1,prelog,indent=indent+subindent)
                i+=1
        data = Data(data_type, values)
        data.drop_all_na()
        if index in data.variables():
            data.set_index(index, print_progress=False)
        return data

    def to_excel(self,path_file:str, sheet:str='sheet1', na:str='',replace_file:bool=False, replace_sheet:bool=False,
                 print_progress:bool=False, indent:int=0, subindent:int=5)->None:
        '''
        write data to a excel file.\n
        'openpyxl' must installed.\n
        pip install openpyxl
        '''
        if print_progress:
            file_name = path_file.split('//')[-1]
            print(' '*indent+f'save data to excel file: {file_name}')
            start, prelog = time.perf_counter(), 0
        import openpyxl
        if os.path.exists(path_file) and not replace_file:
            wb = openpyxl.load_workbook(path_file)
            if replace_sheet and sheet in wb.sheetnames:
                wb.remove(wb.get_sheet_by_name(sheet))
            else:
                if sheet in wb.sheetnames:
                    k = 1
                    while f'{sheet}_{k}' in wb.sheetnames:
                        k += 1
                    sheet = f'{sheet[:26-len(str(k))]}_{k}'
            ws = wb.create_sheet(sheet)
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = sheet
        for index in self.index():
            for j,var in enumerate(self.variables()):
                ws.cell(1,j+2).value = var
        ws.cell(1,1).value = 'index'
        for i,index in enumerate(self.index()):
            for j,var in enumerate(self.variables()):
                ws.cell(i+2,1).value = index
        for i,index in enumerate(indexs:=self.index()):
            for j,var in enumerate(self.variables()):
                if is_nan(val:=self.values[var][index]):
                    ws.cell(i+2,j+2).value = na
                else:
                    ws.cell(i+2,j+2).value = val
            if print_progress:
                prelog = progress(start, i, len(indexs),prelog,index, True, indent+subindent)
        wb.save(path_file)
        return wb
    #endregion
    #region access
    @classmethod
    def read_access(cls, path_file:str, table_name:str, variable_names:list[str]=['*'],
                    data_type:str='cross', index:str='index_', na:str='',
                    print_progress:bool=False, indent:int=0, subindnet:int=5)->Data:
        '''
        read access file.\n
        'pyodbc' must be install.\n
        pip install pyodbc
        '''
        path_file = os.path.join(os.getcwd(), path_file)
        if print_progress:
            file_name = path_file.split('\\')[-1]
            print(' '*indent+f"reading data from {file_name}")
            start_time, prelog = time.perf_counter(), 0
        import pyodbc
        conn = pyodbc.connect(r'Driver={Microsoft Access Driver (*.mdb, *.accdb)};DBQ='+f'{path_file};')
        cursor = conn.cursor()
        if variable_names != ['*'] and not index in variable_names:
            variable_names = variable_names
        cursor.execute(f'select {",".join(variable_names)} from {table_name}')
        variable_names = [column[0] for column in cursor.description]
        if index == '' or not index in variable_names:
            values = {v:{} for v in variable_names}
            records = cursor.fetchall()
            i, n = 0, len(records)
            for vals in records:
                for j, val in enumerate(vals):
                    if val == na:
                        values[variable_names[j]][i] = math.nan
                    else:
                        values[variable_names[j]][i] = val
                if print_progress:
                    prelog = progress(start_time, i, n, prelog, indent=indent+subindnet)
                i += 1
        else:
            values = {v:{} for v in variable_names if v != index}
            records = cursor.fetchall()
            n, i = len(records), 0
            for vals in records:
                for j, val in enumerate(vals):
                    if j != variable_names.index(index):
                        if val == na:
                            values[variable_names[j]][vals[variable_names.index(index)]] = math.nan
                        else:
                            values[variable_names[j]][vals[variable_names.index(index)]] = val
                if print_progress:
                    prelog = progress(start_time, i, n, prelog, indent=indent+subindnet)
                    i += 1
        return cls(data_type, values=values)
    
    def to_access(self, path_file:str, table_name:str, only_new_index:bool=False, na:str='', index:str='index_',
                  print_progress:bool=False, indent:int=0, subindent:int=5)->None:
        '''
        saving data to a access file.\n
        pyodbc and msaccessdb must be installed.
        '''
        path_file = os.path.join(os.getcwd(), path_file)
        file_name = path_file.split('\\')[-1]
        if print_progress:
            print(' '*indent+f"saving data to {file_name}")
            start_time, prelog = time.perf_counter(), 0
        # create access file
        if not os.path.exists(path_file):
            import msaccessdb
            msaccessdb.create(path_file)
        # connction
        import pyodbc
        conn = pyodbc.connect(r'Driver={Microsoft Access Driver (*.mdb, *.accdb)};DBQ='+f'{path_file};')
        cursor = conn.cursor()
        tables = [row.table_name for row in cursor.tables()]
        if not table_name in tables:
            # create table
            sql_statement = f'create table {table_name} (\n'
            sql_statement += f' {index} longtext,\n'
            for var in self.variables():
                for i in self.index():
                    strs, ints, floats, bools = 0, 0, 0, 0
                    if type(val:=self.values[var][i]) == str:
                        strs += 1
                    elif type(val) == int:
                        ints += 1
                    elif type(val) == float:
                        floats += 1
                    elif type(val) == bool:
                        bools += 1
                if max(strs, ints, floats) == strs:
                    var_type = 'char'
                elif max(strs, ints, floats) == ints:
                    var_type = 'int'
                elif max(strs, ints, floats) == floats:
                    var_type = 'decimal'
                elif max(strs, ints, floats) == bools:
                    var_type = 'bool'
                sql_statement += f' {var} {var_type},\n'
            sql_statement = sql_statement[:-2] + '\n);'
            cursor.execute(sql_statement)
            cursor.commit()
            # insert
            indices = self.index()
            n = len(indices)
            for j, i in enumerate(indices):
                sql_statement = f'INSERT INTO {table_name} ('
                for var in [index]+self.variables():
                    sql_statement += f' {var},'
                sql_statement = sql_statement[:-1] + ') VALUES ('
                sql_statement += f" '{i}',"
                for var in self.variables():
                    if isinstance((val:=self.values[var][i]), str):
                        sql_statement += f" '{val}',"
                    else:
                        sql_statement += f" {val},"
                sql_statement = sql_statement[:-1] + ');'
                cursor.execute(sql_statement)
                cursor.commit()
                if print_progress:
                    prelog = progress(start_time, j, n, prelog, i, True, indent+subindent)
        else:
            #update
            sql_statement = f'SELECT index_ from {table_name}'
            cursor.execute(sql_statement)
            old_indices = [row[0] for row in cursor.fetchall()]
            d = 60
            n_parts = int((n:=len(vars:=self.variables()))/d)
            parts = [vars[s*d:(s+1)*d] for s in range(n_parts)]
            parts[-1] += vars[n_parts*d:]
            indices = self.index() if not only_new_index else [i for i in self.index() if not i in old_indices]
            n = len(indices)
            for j, i in enumerate(indices):
                if i in old_indices:
                    for part in parts:
                        sql_statement = f'UPDATE {table_name} SET'
                        for var in part:
                            if isinstance((val:=self.values[var][i]), str):
                                sql_statement += f" {var} = '{val}',"
                            else:
                                sql_statement += f" {var} = {val},"
                        sql_statement = sql_statement[:-1] + f" WHERE {index} = '{i}'"
                        cursor.execute(sql_statement)
                        cursor.commit()
                else:
                    sql_statement = f'INSERT INTO {table_name} ('
                    for var in [index] + self.variables():
                        sql_statement += f" {var},"
                    sql_statement = sql_statement[:-1] + ') VALUES ('
                    sql_statement += f" '{i}',"
                    for var in self.variables():
                        if isinstance((val:=self.values[var][i]), str):
                            sql_statement += f" '{val}',"
                        else:
                            sql_statement += f" {val},"
                    sql_statement = sql_statement[:-1] + f")"
                    cursor.execute(sql_statement)
                    cursor.commit()
                if print_progress:
                    prelog = progress(start_time, j, n, prelog, i, True, indent+subindent)
            pass
    #endregion
    #endregion

class Sample:
    '''
    Sample is a part of a Data like 'train' or 'test'. bigest Sample include all index of Data.\n
    we can 'split' Sample to two part like train and test by random.\n
    and we can weighting 
    '''
    def __init__(self, data: Data, index:list=[], name:str=None, weights:str='1') -> None:
        self.data = data
        if index == []:
            self.index = data.index()
        else:
            self.index = index
        if not set(self.index).issubset(set(data.index())):
            raise ValueError('sample index is not subset of data index')
        self.name = name
        self.weights = weights
        self.stats = Sample.Stats(self)

    def get_data(self, vars:list=[]) -> Data:
        '''
        converting Sample to Data.
        '''
        if vars == []:
            return self.data.select_index(self.index.copy())
        else:
            return self.data.select_variables(vars).select_index(self.index.copy())

    split_methods = ['random', 'start', 'end']
    def split(self, ratio: float, names: list, method: str = 'random') -> list[Sample]:
        '''
        split Sample to two Sample: ratio*n, (1-ratio)*n (n=number of index of Sample).
        '''
        if method == 'random':
            if self.weights == '1':
                S1 = list(np.random.choice(self.index, int(ratio*len(self.index)), replace=False))
            else:
                ws = sum([w for i, w in self.data.values[self.weights].items() if i in self.index])
                weights = [w/ws for i, w in self.data.values[self.weights].items() if i in self.index]
                S1 = list(np.random.choice(self.index, int(ratio*len(self.index)), p=weights, replace=False))

            S2 = list(set(self.index)-set(S1))
        elif method == 'start':
            n = int(ratio * len(self.index))
            S1, S2 = self.index[:n], self.index[n:]
        elif method == 'end':
            n = int((1-ratio) * len(self.index))
            S1, S2 = self.index[:n], self.index[n:]
        return Sample(self.data, S1, names[0], self.weights), Sample(self.data, S2, names[1], self.weights)

    def get_weights(self, path_file_csv:str)-> None:
        '''
        weighting sample with census statistics.\n
        Table of aggregates of society is in a csv file.\n
        for example:\n
        ---------------------------------------------------\n
        |              group                 | population |\n
        ---------------------------------------------------\n
        |      (sex=female)*(age<=18)        |     14     |\n
        |  (sex=female)*(age>18)*(age<=30)   |     31     |\n
        |       (sex=female)*(age>30)        |     26     |\n
        |        (sex=male)*(age<=18)        |     13     |\n
        |    (sex=male)*(age>18)*(age<=30)   |     35     |\n
        |         (sex=male)*(age>30)        |     27     |\n
        ---------------------------------------------------\n
        '''
        # vars_conditions:list[list], totals:list[Union[int,float]]
        groups = Data.read_csv(path_file_csv)
        groups_n = len(groups.index())
        set_index, set_totals = False, False
        for var in groups.variables():
            strs, nums = 0, 0
            for i in groups.index():
                if is_numeric(groups.values[var][i]):
                    nums += 1
                else:
                    strs += 1
            if strs == groups_n:
                groups.set_index(var)
                set_index = True
            elif nums == groups_n:
                totals = list(groups.values[var].values())
                set_totals = True
            else:
                raise ValueError(f"Error! {var} includes numbers and strings, data must be include a string variable as group name and a numeric variable as population of group.")
        if set_index and set_totals:
            vars_conditions = []
            for g in groups.index():
                r = []
                for v in g.split('*'):
                    if len(v.split('>='))>1:
                        i = v.split('>=')
                        sep = '>='
                    elif len(v.split('<='))>1:
                        i = v.split('<=')
                        sep = '<='
                    elif len(v.split('<'))>1:
                        i = v.split('<')
                        sep = '<'
                    elif len(v.split('>'))>1:
                        i = v.split('>')
                        sep = '>'
                    elif len(v.split('='))>1:
                        i = v.split('=')
                        sep = '='
                    if i[1].isdigit():
                        i[1] = int(i[1])
                    elif is_numeric_str(i[1]):
                        i[1] = float(i[1])
                    r.append((i[0],sep,i[1]))
                vars_conditions.append(r)

            new_data=self.data.add_dummies(vars_conditions)
            vars_num = new_data.to_numpy()
            totals_num = np.array(totals)
            I = np.ones(len(self.data.index()))
            D = np.identity(len(self.data.index()))
            w_values_list = list(D @ I + D @ vars_num @ np.linalg.inv(vars_num.T @
                            D @ vars_num) @ np.transpose(totals_num.T - I.T @ D.T @ vars_num))
            w_values = dict(zip(self.data.index(), w_values_list))
            res = {}
            res['weights'] = w_values
            self.data.values.update(res)
            self.weights = 'weights'
        else:
            raise ValueError(f"Error! data must be include a string variable as group name and a numeric variable as population of group.")

    def group(self, by_variable:str, columns:list|str=[], method:str='average',
              filter:list[tuple]=[], filter_operator:str='and',

              print_progress:bool=False, indent:int=0,subindent:int=5)->Data:
        '''
        create a pivot table.\n
        - by_variable: a variable as a pivot of table. (-> note 1)\n
        - columns: or a list of some variables if method is 'average' for all variables, or a list of tuples [('var','method'), ...].\n
        - method: a function, by default 'average'.\n
        - filter: a dict of some variables and their values as a conditions {'var1':'val1', 'var2':'val2', ...}.
        - filter_operator: a function on two or more conditions in filter: 'and' or 'or'.
        \n
        mathods:\n
        - count: the number of non-nan values.\n
        - distinct_count: the number of non-nan and non-repeating values.
        - weight: the summation of sample weights of non-nan values.\n
        - sum: summation.\n
        - average\n
        - var: variance.\n
        - std: standard deviation = variance**0.5.\n
        - min: minimum.\n
        - max: maximum.\n
        Note 1: The pivot variable is usually 'categorical'. If you have a 'numeric' variable, first create a new variable using the 'add_a_group_variable' method on the 'Data' class then use 'group' on the 'Sample' class.\n
        '''
        if columns==[]:
            columns = self.data.variables()
        if isinstance(columns, str):
            columns = [(columns, method)]
        elif not isinstance(columns[0], tuple):
            columns = [(col, method) for col in columns]
        vals, indexs = set(), set()
        for i in self.index:
            filter_cond = True if filter_operator=='and' else False
            by_val = self.data.values[by_variable][i]
            for var_filter, val_filter in filter:
                if not match_str(self.data.values[var_filter][i], val_filter):
                    if filter_operator=='and':
                        filter_cond = False
                        break
                elif filter_operator=='or':
                    filter_cond = True
                    break
            if not is_nan(by_val) and filter_cond:
                vals.add(by_val)
                indexs.add(i)
        vals = sorted(vals)
        values = {f'{var}_{method}':{} for var, method in columns}
        if print_progress:
            print(' '*indent+f'grouping by {by_variable}')
            start, prelog, n, j = time.perf_counter(),0,len(columns), -1
        for var, method in columns:
            counts = {str(val):0 for val in vals}
            distinct_count = {str(val):set() for val in vals}
            if not method in ['count', 'distinct_count']:
                ws = {str(val):0 for val in vals}
                sums = {str(val):0 for val in vals}
                sws = {str(val):0 for val in vals}
                sqs = {str(val):0 for val in vals}
                mins = {str(val):float('+inf') for val in vals}
                maxs = {str(val):float('-inf') for val in vals}
            for i in indexs:
                if not is_nan(x:=self.data.values[var][i]):
                    try:
                        counts[v:=str(self.data.values[by_variable][i])] += 1
                        distinct_count[v].add(x)
                    except Exception as e:
                        try:
                            counts[v:='blank/error'] += 1
                        except:
                            vals.add('blank/error')
                            counts['blank/error'] = 1
                            if not method in ['count', 'distinct_count']:
                                ws['blank/error'] = 0
                                sums['blank/error'] = 0
                                sws['blank/error'] = 0
                                sqs['blank/error'] = 0
                                mins['blank/error'] = 0
                                maxs['blank/error'] = 0
                    if not method in ['count', 'distinct_count'] and not isinstance(x, str):
                        w = 1 if self.weights=='1' else self.data.values[self.weights][i]
                        ws[v] += w
                        sums[v] += x
                        sws[v] += w * x
                        sqs[v] += w * x**2
                        mins[v] = min(mins[v], x)
                        maxs[v] = max(maxs[v], x)
            
            if method == 'count':
                values[f'{var}_{method}'] = counts
            if method == 'distinct_count':
                values[f'{var}_{method}'] = {val:len(distinct_count[val]) for val in distinct_count}
            elif method == 'sum':
                values[f'{var}_{method}'] = sums
            elif method == 'average':
                for val in vals:
                    try:
                        values[f'{var}_{method}'][val] = sws[val]/ws[val] if ws[val]>0 else math.nan
                    except:
                        values[f'{var}_{method}'][str(val)] = sws[str(val)]/ws[str(val)] if ws[str(val)]>0 else math.nan
            elif method == 'var':
                for val in vals:
                    values[f'{var}_{method}'][val] = (sqs[val]-sws[val]**2/ws[val])/ws[val] if ws[val]>0 else math.nan
            elif method == 'std':
                for val in vals:
                    values[f'{var}_{method}'][val] = math.sqrt((sqs[val]-sws[val]**2/ws[val])/ws[val]) if ws[val]>0 else math.nan
            elif method == 'min':
                for val in vals:
                    if mins[val] == float('+inf'):
                        mins[val] = math.nan
                values[f'{var}_{method}'][val] = mins
            elif method == 'max':
                for val in vals:
                    if maxs[val] == float('-inf'):
                        maxs[val] = math.nan
                values[f'{var}_{method}'][val] = mins
            if print_progress:
                prelog = progress(start, j:=j+1, n, prelog, var, indent=indent+subindent)
        return Data(values=values)

    def __len__(self):
        return len(self.index)
    
    def __str__(self) -> str:
        return str(self.get_data())

    class Stats:
        def __init__(self, sample:Sample) -> None:
            self.sample = sample
        
        def count(self, variable:str|list[str]=[],
                   print_progress:bool=False, indent:int=0, subindent:int=5)->int|dict:
            '''count of non-weighted non-nan values of variable in sample.'''
            if not isinstance(variable,list):
                variable = [variable]
            if variable == []:
                variable = self.sample.data.variables()
            if print_progress:
                if len(variable)>1:
                    print(' '*indent+f'count of {variable[0]} ... {variable[-1]}')
                else:
                    print(' '*indent+f'count of {variable[0]}')
                start, prelog, vars_no, k= time.perf_counter(), 0, len(variable), 0
            res = {}
            for j, var in enumerate(variable):
                n = 0
                for i in self.sample.index:
                    if not is_nan(val:=self.sample.data.values[var][i]):
                        n += 1
                res[var] = n
                if print_progress:
                    prelog = progress(start, j, vars_no, prelog, var, True, indent+subindent)
            return res if len(res)>1 else res[variable[0]]

        def weight(self, variable:str|list[str]=[],
                   print_progress:bool=False, indent:int=0, subindent:int=5)->int|dict:
            '''summation of weights of non-nan values of variable in sample.'''
            if not isinstance(variable,list):
                variable = [variable]
            if variable == []:
                for v in self.sample.data.variables():
                    try:
                        if self.sample.data.dtype(v) == 'numeric':
                            variable.append(v)
                    except:
                        pass
            if print_progress:
                if len(variable)>1:
                    print(' '*indent+f'count of {variable[0]} ... {variable[-1]}')
                else:
                    print(' '*indent+f'count of {variable[0]}')
                start, prelog, vars_no, k= time.perf_counter(), 0, len(variable), 0
            res = {}
            for j, var in enumerate(variable):
                n = 0
                for i in self.sample.index:
                    if not is_nan(val:=self.sample.data.values[var][i]):
                        w = 1 if self.sample.weights=='1' else self.sample.data.values[self.sample.weights][i]
                        n += w
                res[var] = w
                if print_progress:
                    prelog = progress(start, j, vars_no, prelog, var, True, indent+subindent)
            return res if len(res)>1 else res[variable[0]]

        def sum(self, variable:str|list[str]=[], weighted:bool=False, print_progress:bool=False, indent:int=0, subindent:int=5)->float|dict:
            '''summation of non-weighted or weighted non-nan values.\n
            - variable: one or more variables.\n
            - weighted: True for summation of products of weights and non-values. False for summation of only non-nan values. default = False.\n
            '''
            if not isinstance(variable,list):
                variable = [variable]
            if variable == []:
                for v in self.sample.data.variables():
                    try:
                        if self.sample.data.dtype(v) == 'numeric':
                            variable.append(v)
                    except:
                        pass
            if print_progress:
                if len(variable)>1:
                    print(' '*indent+f'sum of {variable[0]} ... {variable[-1]}')
                else:
                    print(' '*indent+f'sum of {variable[0]}')
                start, prelog, vars_no, k= time.perf_counter(), 0, len(variable), 0
            res = {}
            for j, var in enumerate(variable):
                s, ws = 0, 0
                for i in self.sample.index:
                    if not is_nan(val:=self.sample.data.values[var][i]):
                        w = 1 if self.sample.weights=='1' else self.sample.data.values[self.sample.weights][i]
                        s += val
                        ws += w * val
                if weighted:
                    res[var] = ws
                else:
                    res[var] = s
                if print_progress:
                    prelog = progress(start, j, vars_no, prelog, var, True, indent+subindent)
            return res if len(res)>1 else res[variable[0]]

        def average(self, variable:str|list[str]=[], weighted:bool=True, print_progress:bool=False, indent:int=0, subindent:int=5)->float|dict:
            '''average of non-weighted or weighted non-nan values.\n
            - variable: one or more variables.\n
            - weighted: True for average of non-values. False for summation of only non-nan values. default = True.\n'''
            if not isinstance(variable,list):
                variable = [variable]
            if variable == []:
                for v in self.sample.data.variables():
                    try:
                        if self.sample.data.dtype(v) == 'numeric':
                            variable.append(v)
                    except:
                        pass
            if print_progress:
                if len(variable)>1:
                    print(' '*indent+f'average of {variable[0]} ... {variable[-1]}')
                else:
                    print(' '*indent+f'average of {variable[0]}')
                start, prelog, vars_no, k= time.perf_counter(), 0, len(variable), 0
            res = {}
            for j, var in enumerate(variable):
                n, ws, s, sw = 0, 0, 0, 0
                for i in self.sample.index:
                    if not is_nan(val:=self.sample.data.values[var][i]):
                        w = 1 if self.sample.weights=='1' else self.sample.data.values[self.sample.weights][i]
                        n += 1
                        ws += w
                        s += val
                        sw += val * w
                if n>0:
                    if weighted:
                        res[var] = s/n
                    else:
                        res[var] = sw/ws
                else:
                    res[var] = math.nan
                if print_progress:
                    prelog = progress(start, j, vars_no, prelog, var, True, indent+subindent)
            return res if len(res)>1 else res[variable[0]]

        def var(self, variable:str|list[str]=[], weighted:bool=True, print_progress:bool=False, indent:int=0, subindent:int=5)->float|dict:
            '''variance of non-weighted or weighted non-nan values.\n
            - variable: one or more variables.\n
            - weighted: True for average of non-values. False for summation of only non-nan values. default = True.\n'''
            if not isinstance(variable,list):
                variable = [variable]
            if variable == []:
                for v in self.sample.data.variables():
                    try:
                        if self.sample.data.dtype(v) == 'numeric':
                            variable.append(v)
                    except:
                        pass
            if print_progress:
                if len(variable)>1:
                    print(' '*indent+f'variance of {variable[0]} ... {variable[-1]}')
                else:
                    print(' '*indent+f'variance of {variable[0]}')
                start, prelog, vars_no, k= time.perf_counter(), 0, len(variable), 0
            res = {v:{} for v in variable}
            for j, var1 in enumerate(variable):
                for var2 in variable:
                    n, s1, s2, s12, s1_2, s2_2 = 0, 0, 0, 0, 0, 0
                    ws, sw1, sw2, sw12, sw1_2, sw2_2 = 0, 0, 0, 0, 0, 0
                    for i in self.sample.index:
                        if not is_nan(val1:=self.sample.data.values[var1][i]) and \
                            not is_nan(val2:=self.sample.data.values[var2][i]):
                            w = 1 if self.sample.weights=='1' else self.sample.data.values[self.sample.weights][i]
                            n += 1
                            s1 += val1
                            s2 += val2
                            s12 += val1*val2
                            ws += w
                            sw1 += val1 * w
                            sw2 += val2 * w
                            sw12 += val1*val2 * w
                    if n>0:
                        if weighted:
                            res[var1][var2] = (s12-s1*s2/n)/n
                        else:
                            res[var1][var2] = (sw12-sw1*sw2/ws)/ws
                    else:
                        res[var1][var2] = math.nan
                if print_progress:
                    k += vars_no -j if vars_no>1 else vars_no -j-1
                    prelog = progress(start, k, vars_no*(vars_no+1)//2, prelog, var1, True, indent+subindent)
            return res if len(variable) >1 else res[variable[0]][variable[0]]

        def std(self, variable:str|list[str]=[], weighted:bool=True, print_progress:bool=False, indent:int=0, subindent:int=5)->float|dict:
            '''standard deviation (sqart of variance) of non-weighted or weighted non-nan values.\n
            - variable: one or more variables.\n
            - weighted: True for average of non-values. False for summation of only non-nan values. default = True.\n'''
            vars = self.var(variable, weighted, print_progress, indent, subindent)
            if not isinstance(variable, list):
                return vars**.5
            for var1 in variable:
                for var2 in variable:
                    vars[var1][var2] **= .5
            return vars

        def distribution(self, variable:str, weighted:bool=True, print_progress:bool=False, indent:int=0, subindent:int=5) -> dict:
            '''count and weight of values of a variable in sample as a dict: {'var':{'count':{}, 'weight':{}}}.\n
            if weighted=True(default): weight = sum of weights of non-nan values of a group divided by total weights.\n
            if weighted=False: weight = counts of non-nan values divided by total counts.\n'''
            if print_progress:
                if isinstance(variable, list):
                    print(' '*indent+f"distribution of {variable[0]}...{variable[-1]}")
                else:
                    print(' '*indent+f"distribution of {variable}")
                start, prelog, counts = time.perf_counter(), 0, len(self.sample.index)
            ws, freq, wgh, n = 0, {}, {}, 0
            for i in self.sample.index:
                if not is_nan(x:=self.sample.data.values[variable][i]):
                    w = 1 if self.sample.weights == '1' else self.sample.data.values[self.sample.weights][i]
                    n += 1
                    ws += w
                    if x in freq.keys():
                        freq[x] += 1
                        wgh[x] += w
                    else:
                        freq[x], wgh[x] = 1, w
                if print_progress:
                    prelog = progress(start, i, counts, prelog, indent=indent+subindent)
            if ws > 0:
                dist = {}
                if weighted:
                    for v, f in freq.items():
                        dist[v] = {'count': f, 'weight': wgh[v]/ws}
                else:
                    for v, f in freq.items():
                        dist[v] = {'count': f, 'weight': f/n}
                return dist

        def median(self, variable:str|list[str]=[],weighted:bool=True, print_progress:bool=False, indent:int=0, subindent:int=5)->float|dict:
            '''median (percentile of 0.5) of non-weighted or weighted non-nan values.\n
            - variable: one or more variables.\n
            - weighted: default = True.\n'''
            if print_progress:
                if not isinstance(variable, list):
                    print(' '*indent+f'average of {variable}')
                else:
                    print(' '*indent+f'average of {variable[0]} ... {variable[-1]}')
            if not isinstance(variable,list):
                variable = [variable]
            if variable == []:
                for v in self.sample.data.variables():
                    try:
                        if self.sample.data.dtype(v) == 'numeric':
                            variable.append(v)
                    except:
                        pass
            res = {}
            start, prelog, vars_no, k= time.perf_counter(), 0, len(variable), 0
            for j, var in enumerate(variable):
                F2 = 0
                for v2 in sorted(dist:=self.distribution(var, weighted)):
                    F2 += (f2:=dist[v2]['weight'])
                    if F2 == 0.5:
                        res[var] = v2
                        break
                    elif F2>0.5:
                        res[var] = (v1+v2)/2+(v2-v1)/f2*(1-F1-F2)/2
                        break
                    v1, F1 = v2, F2
                if print_progress:
                    prelog = progress(start, j, vars_no, prelog, indent=indent+subindent)
            if len(variable)==1:
                return res[variable[0]]
            return res

        def mode(self, variable:str|list[str]=[],weighted:bool=True, print_progress:bool=False, indent:int=0, subindent:int=5):
            '''the value of maximum frequency.'''
            if print_progress:
                if not isinstance(variable, list):
                    print(' '*indent+f'mode of {variable}')
                else:
                    print(' '*indent+f'mode of {variable[0]} ... {variable[-1]}')
            if not isinstance(variable,list):
                variable = [variable]
            if variable == []:
                for v in self.sample.data.variables():
                    try:
                        if self.sample.data.dtype(v) == 'numeric':
                            variable.append(v)
                    except:
                        pass
            res = {}
            start, prelog, vars_no, k= time.perf_counter(), 0, len(variable), 0
            for j, var in enumerate(variable):
                dist = self.distribution(var, weighted)
                max_w = max(dist[val]['weight'] for val in dist)
                for val in dist:
                    w = dist[val]['weight']
                    if w == max_w:
                        res[var] = val
                if print_progress:
                    prelog = progress(start, j, vars_no, prelog, indent=indent+subindent)
            if len(variable)==1:
                return res[variable[0]]
            return res

        def correl(self, variables:list[str]=[], weighted:bool=True, print_progress:bool=False, indent:int=0, subindent:int=5)->dict:
            '''pair correlations variables. output is a dictionary: {'var1':{'var1':correl, 'var2':correl,...}, 'var2':{'var1':correl, 'var2':correl,...},... }.\n
            you can put output on a Data: Data(values=correl()).\n
            - variables: a list of variables include at least two variable.\n
            - weighted: True for average of non-values. False for summation of only non-nan values. default = True.\n'''
            # print(self.sample.data.variables())
            if variables == []:
                for v in self.sample.data.variables():
                    try:
                        if self.sample.data.dtype(v) == 'numeric':
                            variables.append(v)
                    except:
                        pass
            if print_progress:
                print(' '*indent+f'correl of {variables[0]} ... {variables[-1]}')
            if len(variables)==1:
                raise ValueError("variables must be more than one.")
            ind = {}
            for var in variables:
                ind[var] = {i for i in self.sample.index if not is_nan(self.sample.data.values[var][i])}
            res = {v1:{v2:1 if v2==v1 else math.nan for v2 in variables} for v1 in variables}
            if print_progress:
                start, prelog, m, k= time.perf_counter(), 0, sum([len(ind[v]) for v in variables]), 0
            for j, var1 in enumerate(variables[:-1]):
                n, s1, s2, s12, s1_2, s2_2 = 0, 0, 0, 0, 0, 0
                for var2 in variables[j+1:]:
                    for i in (ind[var1].intersection(ind[var2])):
                        w = 1 if self.sample.weights=='1' else self.sample.data.values[self.sample.weights][i]
                        n += w if weighted else 1
                        val1=self.sample.data.values[var1][i]
                        val2=self.sample.data.values[var2][i]
                        s1 += val1*w if weighted else val1
                        s2 += val2*w if weighted else val2
                        s12 += val1*val2*w if weighted else val1*val2
                        s1_2 += val1**2*w if weighted else val1**2
                        s2_2 += val2**2*w if weighted else val2**2
                    try:
                        res[var1][var2] = (s12-s1*s2/n)/math.sqrt((s1_2-s1**2/n)*(s2_2-s2**2/n))
                    except:
                        pass
                    res[var2][var1] = res[var1][var2]
                if print_progress:
                    k += len(ind[var1])
                    prelog = progress(start, k, m, prelog, var1, True, indent+subindent)
            return res

        def min(self, variable:str|list[str]=[],
                   print_progress:bool=False, indent:int=0, subindent:int=5)->int|dict:
            '''minimum of non-nan values of one or more variables on sample. output is float if a variabe otherwise is a dict: {'var':'min', ...}'''
            if not isinstance(variable,list):
                variable = [variable]
            if variable == []:
                for v in self.sample.data.variables():
                    try:
                        if self.sample.data.dtype(v) == 'numeric':
                            variable.append(v)
                    except:
                        pass
            if print_progress:
                if len(variable)>1:
                    print(' '*indent+f'min of {variable[0]} ... {variable[-1]}')
                else:
                    print(' '*indent+f'min of {variable[0]}')
                start, prelog, vars_no, k= time.perf_counter(), 0, len(variable), 0
            res = {}
            for j, var in enumerate(variable):
                min_val = float('+inf')
                for i in self.sample.index:
                    if not is_nan(val:=self.sample.data.values[var][i]):
                        if min_val > val:
                            min_val = val
                res[var] = min_val
                if print_progress:
                    prelog = progress(start, j, vars_no, prelog, var, True, indent+subindent)
            return res if len(res)>1 else res[variable[0]]

        def max(self, variable:str|list[str]=[],
                   print_progress:bool=False, indent:int=0, subindent:int=5)->int|dict:
            '''maximum of non-nan values of one or more variables on sample. output is float if a variabe otherwise is a dict: {'var':'min', ...}'''
            if not isinstance(variable,list):
                variable = [variable]
            if variable == []:
                for v in self.sample.data.variables():
                    try:
                        if self.sample.data.dtype(v) == 'numeric':
                            variable.append(v)
                    except:
                        pass
            if print_progress:
                if len(variable)>1:
                    print(' '*indent+f'max of {variable[0]} ... {variable[-1]}')
                else:
                    print(' '*indent+f'max of {variable[0]}')
                start, prelog, vars_no, k= time.perf_counter(), 0, len(variable), 0
            res = {}
            for j, var in enumerate(variable):
                max_val = float('-inf')
                for i in self.sample.index:
                    if not is_nan(val:=self.sample.data.values[var][i]):
                        if max_val < val:
                            max_val = val
                res[var] = max_val
                if print_progress:
                    prelog = progress(start, j, vars_no, prelog, var, True, indent+subindent)
            return res if len(res)>1 else res[variable[0]]

        def percentile(self, p:float, variable:str|list[str]=[],weighted:bool=True, print_progress:bool=False, indent:int=0, subindent:int=5)->float|dict:
            '''percentile of non-weighted or weighted non-nan values.\n
            - variable: one or more variables.\n
            - weighted: default = True.\n'''
            if print_progress:
                if not isinstance(variable, list):
                    print(' '*indent+f'average of {variable}')
                else:
                    print(' '*indent+f'average of {variable[0]} ... {variable[-1]}')
            if not isinstance(variable,list):
                variable = [variable]
            if variable == []:
                for v in self.sample.data.variables():
                    try:
                        if self.sample.data.dtype(v) == 'numeric':
                            variable.append(v)
                    except:
                        pass
            res = {}
            start, prelog, vars_no, k= time.perf_counter(), 0, len(variable), 0
            for j, var in enumerate(variable):
                F2 = 0
                for v2 in sorted(dist:=self.distribution(var, weighted)):
                    F2 += (f2:=dist[v2]['weight'])
                    if F2 == p:
                        res[var] = v2
                        break
                    elif F2 > p:
                        res[var] = (v1+v2)/2+(v2-v1)/f2*(2*p-F1-F2)/2
                        break
                    v1, F1 = v2, F2
                if print_progress:
                    prelog = progress(start, j, vars_no, prelog, indent=indent+subindent)
            if len(variable)==1:
                return res[variable[0]]
            return res

        def gini(self, variable:str|list[str]=[],weighted:bool=True, print_progress:bool=False, indent:int=0, subindent:int=5)->float|dict:
            '''gini coeficient is a measure of statistical dispersion intended to represent the income inequality, the wealth inequality, or the consumption inequality.\n
            we used Discrete probability distribution formula. -> https://en.wikipedia.org/wiki/Gini_coefficient.\n
            - variable: one or more variables.\n
            - weighted: default = True.\n'''
            if print_progress:
                if not isinstance(variable, list):
                    print(' '*indent+f'gini of {variable}')
                else:
                    print(' '*indent+f'gini of {variable[0]} ... {variable[-1]}')
            if not isinstance(variable,list):
                variable = [variable]
            if variable == []:
                for v in self.sample.data.variables():
                    try:
                        if self.sample.data.dtype(v) == 'numeric':
                            variable.append(v)
                    except:
                        pass
            res = {}
            start, prelog, vars_no, k= time.perf_counter(), 0, len(variable), 0
            for j, var in enumerate(variable):
                ys = sorted(dist:=self.distribution(var, weighted))
                H, Si, Si_1 = 0, 0, 0
                for y in ys:
                    Si += dist[y]['weight']*y
                    H += dist[y]['weight']*(Si_1+Si)
                    Si_1 = Si
                res[var] = 1-H/Si # Si==Sn
                if print_progress:
                    prelog = progress(start, j, vars_no, prelog, var, indent=indent+subindent)
            return res

class TimeSeries(Data):
    '''
    TimeSeries is child of Data for developing 'time' type of Data.
    '''
    types = ['daily', 'weekly', 'monthly', 'seasonal', 'annual']
    @staticmethod
    def str_to_date(date:str)->jdatetime.date:
        '''
        convert str (1402-04-02) to jdatetime.date.
        '''
        seprator = '-' if len(date.split('-')) != 1 else '/' if len(date.split('/')) != 1 else ''

        if seprator == '':
            raise ValueError(f"Error! seprator are not standard: '-' or '/'.")
        if len(date.split(seprator))!=3:
            raise ValueError(f"Error! date must have three parts: year, month and day.")
        return jdatetime.date(*[int(p) for p in date.split(seprator)])

    @staticmethod
    def type_of_dates(dates:list[str]):
        '''
        recognizing type of date: annual, seasonal, monthly, weekly or daily base on values of index.
        '''
        if dates == []:
            return 'daily'
        elif is_numeric(dates[0]):
            return 'annual'
        else:
            seprator = '-' if len(dates[0].split('-')) != 1 else '/' if len(dates[0].split('/')) != 1 else ''
            if seprator == '':
                raise ValueError(f"Error! seprator are not standard: '-' or '/'.")
            splits = dates[0].split(seprator)
            if len(splits) == 2:
                try:
                    parts2 = [int(date.split(seprator)[1]) for date in dates]
                    if max(parts2) == 4:
                        return 'seasonal'
                    else:
                        return 'monthly'
                except:
                    raise ValueError(f"Error! dates have two parts (they are monthly or seasonal), but probably one date is different from other dates., it has one or more than two parts.")
            elif len(splits) == 3:
                try:
                    sorted_dates = []
                    for date in sorted(dates):
                        seprator = '-' if len(date.split('-')) != 1 else '/' if len(date.split('/')) != 1 else ''
                        if seprator == '':
                            raise ValueError(f"Error! date {date} seprator are not standard: '-' or '/'.")
                        sorted_dates.append(jdatetime.date(*[int(p) for p in date.split(seprator)]))
                    difs = set([(sorted_dates[i]-sorted_dates[i-1]).days for i in range(1,len(sorted_dates))])
                    weekly = True
                    for dif in difs:
                        if dif % 7 != 0:
                            weekly = False
                            break
                    if weekly:
                        return 'weekly'
                    else:
                        return 'daily'
                except:
                    raise ValueError(f"Error! dates have three parts (they are daily or weekly), but probably one date is different from other dates, it has one or two than two parts.")
            else:
                raise ValueError(f"Error! the number of parts '{dates[0]}' must be 2 (seasonal or monthly) or 3 (daily or weekly).")

    def __init__(self, type:str='time', values: dict[dict] = {}) -> None:
        self.type = type
        if values != {}:
            self.dates = list(values[list(values.keys())[0]].keys())
        else:
            self.dates = []
        
        self.values = values
        try:
            self.date_type = TimeSeries.type_of_dates(self.dates)
        except:
            self.complete_dates()
            self.date_type  = TimeSeries.type_of_dates(self.dates)
        # super().__init__('time', self.values)

    def complete_dates(self, print_progress:bool=False, indent:int=0, subindent:int=5):
        '''
        recognizing type of date and insert missing index base on type.\n
        for example between 1402-04-28 (wednesday) and 1402-4-31 (saturday) there are 1402-04-29 (thursday) and 1402-04-30 (friday).\n
        '''
        
        if print_progress:
            print(' '*indent+'completing dates...')
            start_time, prelog, n = time.perf_counter(), 0, len(self.dates)
        if self.dates != []:
            if print_progress:
                print(' '*(indent+subindent)+'standardizing dates..')
            is_standard = True
            for i, date in enumerate(self.dates):
                if is_numeric(date):
                    break   # dates are annual
                seprator = '/' if '/' in date else '-' if '-' in date else ''
                if seprator == '':
                    raise ValueError(f"Error! seprator must be '-' or '/'.")
                if self.date_type in ['daily', 'weekly', 'monthly']:
                    sp = [f'0{int(x)}' if int(x) < 10 else f'{int(x)}' 
                            for x in date.split(seprator)]
                    date = '-'.join(sp)
                else:
                    date.replace(seprator, '-')
                if self.dates[i] != date:
                    self.dates[i] = date
                    is_standard = False
                if print_progress:
                    prelog = progress(start_time, i, n, prelog, '', True, indent+2*subindent)
            if not is_standard:
                self.set_index(var=self.dates,drop_var=False, print_progress=print_progress, indent=indent+subindent)
            self.dates.sort()
            start_dates, end_dates = self.dates[0], self.dates[-1]
            if print_progress:
                print(' '*(indent+subindent)+'chacking and completing dates')
                start_time, prelog = time.perf_counter(), 0
            if self.date_type in ['daily', 'weekly']:
                delta = 1 if self.date_type == 'daily' else 7
                start_dates, end_dates = [TimeSeries.str_to_date(date) for date in sorted([start_dates, end_dates])]
                date = start_dates
                n, t = (end_dates-start_dates).days+1, 0
                while (end_dates-date).days>=0:
                    new_date = date.strftime('%Y-%m-%d')
                    if not new_date in self.dates:
                        for var in self.variables():
                            self.values[var][new_date] = math.nan
                        self.dates.append(new_date)
                    date = date + jdatetime.timedelta(delta)
                    if print_progress:
                        prelog = progress(start_time, t, n, prelog, date, True, indent+2*subindent)
                        t+=1
                self.sort(print_progress=False)
            elif self.date_type in ['monthly', 'seasonal']:
                end_month = 13 if self.date_type=='monthly' else 5
                ystart, mstart = [int(p) for p in start_dates.split('-')]
                yend, mend = [int(p) for p in end_dates.split('-')]
                dates = self.dates.copy()
                t, n = 0, yend+1-ystart
                for y in range(ystart, yend+1):
                    st = mstart if y == ystart else 1
                    en = mend+1 if y==yend else end_month
                    for m in range(st, en):
                        mstr = f'{m}' if m>9 else f'0{m}' if self.date_type=='monthly' else f'{m}'
                        new_date = f'{y}-{mstr}'
                        for var in self.values.keys():
                            if not new_date in self.dates:
                                self.values[var][new_date] = math.nan
                        if not new_date in dates:
                            dates.append(new_date)
                    if print_progress:
                        prelog = progress(start_time, t, n, prelog, y, True, indent+2*subindent)
                        t += 1
                self.dates = dates
                self.sort(print_progress=False)
            elif self.date_type == 'annual':
                if not (is_numeric(start_dates) or is_numeric(end_dates)):
                    raise ValueError(f"Error! annual dates must be numeric.")
                y = start_dates
                t, n = 0, end_dates+1-start_dates
                while y <= end_dates:
                    for var in self.values.keys():
                        if not y in self.dates:
                            self.values[var][y] = math.nan
                    if not y in self.dates:
                        self.dates.append(y)
                    y += 1
                    if print_progress:
                        prelog = progress(start_time, t, n, prelog, str(y), True, indent+2*subindent)
                        t += 1
                self.sort(print_progress=False)
            else:
                raise ValueError(
                    f"Error! date_type '{self.date_type}' most be daily, weekly, monthly, seasonal, or annual")

    def reset_date_type(self):
        '''
        setting type of data by use of type_of_dates function base on data. 
        '''
        self.dates = self.index()
        self.date_type = TimeSeries.type_of_dates(self.dates)

    def set_index(self, var: str | list, drop_var: bool = True, print_progress:bool=False, indent:int=0) -> None:
        '''
        setting index base on a variable or list of index out of data
        '''
        super().set_index(var, drop_var, '', print_progress, indent)
        self.reset_date_type()
        self.complete_dates()

    monthly_methods = ['average', 'sum', 'curve']
    def to_monthly(self, method:str, farvardin_adj:bool=False,
                   skip_editing_last_month:bool=False,
                   max_mag_for_editing_last_month:int=3,
                   print_progress:bool=False, indent:int=0, subindent:int=5)->TimeSeries:
        '''
        convert data to a monthly data
        '''
        if print_progress:
            print(' '*indent + 'converting data to monthly.')
        if self.date_type == 'daily':
            #region sums and counts for months
            if print_progress:
                print(' '*(indent+subindent)+'calculation of sums and numbers:')
                start_time, prelog = time.perf_counter(), 0
            sums, counts = {}, {}
            n = len(self.variables())
            sep = '-' if '-' in self.dates[0] else '/' if '/' in self.dates[0] else ''
            if sep == '':
                raise ValueError(
                    f"Error! {self.dates[0]} in variable {var} isn't standard. seprator must be '-' or '/'.")
            if len(self.dates[0].split(sep)) != 3:
                raise ValueError(f"Error! {day} in variable {var} isn't standard. day must has 3 part.")
            for j, var in enumerate(self.variables()):
                sums[var], counts[var] = {}, {}
                for day in self.dates:
                    y, m = [int(x) for x in day.split(sep)[:2]]
                    month = f'{y}-{m}' if m>9 else f'{y}-0{m}'
                    if not is_nan((v:=self.values[var][day]), is_number=True):
                        if month in sums[var].keys():
                            sums[var][month] += v
                            counts[var][month] += 1
                        else:
                            sums[var][month] = v
                            counts[var][month] = 1
                if print_progress:
                    prelog = progress(start_time, j, n, prelog, var, True, indent+2*subindent)
            #endregion
            #region average
            if method == 'average':
                if print_progress:
                    print(' '*(indent+subindent) + 'calculation of averages:')
                    start_time, prelog = time.perf_counter(), 0
                start = True
                for j, var in enumerate(self.variables()):
                    for month in sums[var].keys():
                        if start:
                            month_start, month_end, start = month, month, False
                        else:
                            month_start, month_end = min(month_start, month), max(month_end, month)
                        sums[var][month] /= counts[var][month]
                    if print_progress:
                        prelog = progress(start_time, j, n, prelog, var, True, indent+2*subindent)
            #endregion
            #region complete keys
            def next_month(month:str)->str:
                sep = '-' if '-' in month else '/' if '/' in month else ''
                if sep == '':
                    raise ValueError(
                        f"Error! {month} isn't standard. seprator must be '-' or '/'.")
                if len(month.split(sep)) != 2:
                    raise ValueError(f"Error! {month} isn't standard. month must has 2 part.")
                y, m = [int(x) for x in month.split(sep)]
                return f'{y+1}-01' if m==12 else f'{y}-{m+1}' if m+1>9 else f'{y}-0{m+1}'
            
            month = month_start
            while month <= month_end:
                for var in self.variables():
                    if not month in sums[var].keys():
                        sums[var][month] = math.nan
                month = next_month(month)
            #endregion
            
            res = TimeSeries('time', sums)
            # res.addna()
            res.sort(print_progress=False)

            #region edit last month
            if not skip_editing_last_month:
                if print_progress:
                    print(' '*(indent+subindent)+'calculate edit last month...')
                    start_time, prelog = time.perf_counter(), 0
                for j,var in enumerate(self.variables()):
                    #region end month of var
                    for month in res.dates[::-1]:
                        v = res.values[var][month]
                        if not is_nan(v, is_number=True):
                            month_end = month               # myabe end moth of var don't equal with end month of data
                            break
                    sep = '-' if '-' in month else '/'
                    y, m = [int(x) for x in month_end.split(sep)]
                    #endregion
                    #region number of end month
                    if m == 12:
                        no_days = (jdatetime.date(y+1, 1, 1) - jdatetime.date(y, m, 1)).days
                    else:
                        no_days = 31 if m<7 else 30
                    #endregion
                    #region sum and count of values on end month
                    s1, n1 = 0, 0
                    for d in range(1, no_days+1):
                        m_str = f'{m}' if m>9 else f'0{m}'
                        d_str = f'{d}' if d>9 else f'0{d}'
                        day = f'{y}-{m_str}-{d_str}'
                        if day in self.dates:
                            v = self.values[var][day]
                            if not is_nan(v, is_number=True):
                                s1 += v
                                n1 += 1
                    #endregion
                    #region sum and count of values on month before end month
                    s0 = 0
                    lag = 1     # month before end month
                    while s0 == 0 and lag<=max_mag_for_editing_last_month:
                        s0, n0 = 0, 0
                        for d in range(1, no_days+1):
                            d_str = f'{d}' if d>9 else f'0{d}'
                            o = y*12+m-lag
                            y_lag, m_lag = o//12, o - (o//12)*12
                            m_str0 = f'{m_lag}' if m_lag > 9 else f'0{m_lag}'
                            day0 = f'{y_lag}-{m_str0}-{d_str}'
                            month_lag = f'{y_lag}-{m_str0}'
                            if day0 in self.dates:
                                v0 = self.values[var][day0]
                                if not is_nan(v0, is_number=True):
                                    s0 += v0
                                    n0 += 1
                        lag += 1
                    #endregion
                    if s0 != 0:
                        if method=='average':
                            res.values[var][month_end] = (s1/n1)/(s0/n0) * res.values[var][month_lag]
                        else:
                            res.values[var][month_end] = (s1/s0) * res.values[var][month_lag]
                    if print_progress:
                        prelog = progress(start_time, j, n, prelog, var, True, indent+2*subindent)
                    # else:
                    #     res.values[var][month_end] = math.nan
            #endregion
            return res
        elif self.date_type == 'weekly':
            if print_progress:
                start_time, prelog, n = time.perf_counter(), 0, len(self.variables())
            values = {}
            for var in self.variables():
                values[var] = {}
                last_month = ''
                for week in self.dates:
                    seprator = '-' if len(week.split('-')) != 1 else '/' if len(week.split('/')) != 1 else ''
                    if seprator == '':
                        raise ValueError(f"Error! seprator are not standard: '-' or '/'.")
                    if len(week.split(seprator))==3:
                        y,m,d = [int(x) for x in week.split(seprator)]
                        month = f'{y}-{m}' if m > 9 else f'{y}-0{m}'
                        if month != last_month:
                            summation, last_value, count = 0, math.nan, 0
                        last_month = month
                        value = self.values[var][week]
                        is_nan_value = False
                        if is_numeric(value):
                            if math.isnan(value):
                                is_nan_value = True
                        if not is_nan_value:
                            summation += value
                            count += 1
                            last_value = value
                        else:
                            is_nan_last_value = False
                            if is_numeric(last_value):
                                if math.isnan(last_value):
                                    is_nan_last_value = True
                            if not is_nan_last_value:
                                summation += last_value
                                count += 1
                        if count>0:
                            values[var][month] = summation / count if method=='average' else summation
                        elif not month in values[var].keys():
                            values[var][month] = math.nan
                    else:
                        raise ValueError(f"Error! '{week}' is not a standard format of daily date like yyyy-mm-dd.")
                if print_progress:
                    prelog = progress(start_time, j ,n, var, True, indent+2*subindent)
            res = TimeSeries('time', values)
            res.reset_date_type()
            res.complete_dates(print_progress=print_progress, indent=indent+subindent, subindent=subindent)
            return res
        elif self.date_type == 'monthly':
            return self
        elif self.date_type == 'seasonal':
            if print_progress:
                start_time, prelog, vars_no = time.perf_counter(), 0, len(self.variables())
            res = Data(values={})
            if method == 'curve':
                #region Type Error
                if self.date_type != 'seasonal':
                    raise ValueError(f"Error! 'curve' method only work on 'seanonal' data types.")
                #endregion
                values = {}
                for var_j, var in enumerate(self.variables()):
                    #region start and end values
                    start, nans, dates, month_dates = True, 0, [], []
                    for date in self.dates:
                        # print(var, date, self.values[var][date])
                        if not is_nan(self.values[var][date]):
                            dates.append(date)
                            start = False
                        elif not math.isnan(self.values[var][date]) and not start:
                            if nans > 0:
                                raise ValueError(f"Error! there is a 'nan' bettween values of '{var}'.")
                            n += 1
                            dates.append(date)
                        elif math.isnan(self.values[var][date]) and not start:
                            nans += 1
                    #endregion

                    rows = len(dates) * 3
                    coefs_arr, const_vec = np.zeros((rows, rows)), np.zeros((rows, 1))

                    #region values
                    for i in range(rows):
                        year,season = [int(x) for x in dates[i//3].split('-')]
                        month = (season-1)*3 + i%3 + 1
                        date = f'{year}-0{month}' if month < 10 else f'{year}-{month}'
                        month_dates.append(date)
                        if i == 0:
                            coefs_arr[i][0] = 1             # j=
                            coefs_arr[i][1] = 1             # j=1
                            coefs_arr[i][2] = 1             # j=2
                            coefs_arr[i][3] = 0             # j=3
                            for j in range(4, rows):        # j>3
                                coefs_arr[i][j] = 0
                            const_vec[i][0] = self.values[var][dates[0]]
                        elif i == 1:
                            coefs_arr[i][0] = 2             # j=0
                            coefs_arr[i][1] = -3            # j=1
                            coefs_arr[i][2] = 0             # j=2
                            coefs_arr[i][3] = 1             # j=3
                            for j in range(4, rows):        # j>3
                                coefs_arr[i][j] = 0
                            const_vec[i][0] = 0
                        elif i == 2:
                            coefs_arr[i][0] = 1             # j=0
                            coefs_arr[i][1] = 0             # j=1
                            coefs_arr[i][2] = -3            # j=2
                            coefs_arr[i][3] = 2             # j=3
                            for j in range(4, rows):        # j>3
                                coefs_arr[i][j] = 0
                            const_vec[i][0] = 0
                        elif 2<i<rows-3:
                            for j in range(i-1):            # j<i-1
                                coefs_arr[i][j] = 0
                            if i % 3 == 0:
                                coefs_arr[i][i-1] = 1        # j=i-1
                                coefs_arr[i][i] = -2         # j=i
                                coefs_arr[i][i+1] = 1        # j=i+1
                                coefs_arr[i][i+2] = 0        # j=i+2
                                coefs_arr[i][i+3] = 0        # j=i+3
                                const_vec[i][0] = 0
                            elif i % 3 == 1:
                                coefs_arr[i][i-2] = 0        # j=i-2
                                coefs_arr[i][i-1] = 0        # j=i-1
                                coefs_arr[i][i] = 1          # j=i
                                coefs_arr[i][i+1] = -2       # j=i+1
                                coefs_arr[i][i+2] = 1        # j=i+2
                                const_vec[i][0] = 0
                            elif i % 3 == 2:
                                coefs_arr[i][i-3] = 0        # j=i-3
                                coefs_arr[i][i-2] = 1        # j=i-2
                                coefs_arr[i][i-1] = 1        # j=i-1
                                coefs_arr[i][i] = 1          # j=i
                                coefs_arr[i][i+1] = 0        # j=i+1
                                const_vec[i][0] = self.values[var][dates[i//3]]
                            for j in range(i+4,rows):       # j>i+3
                                coefs_arr[i][j] = 0
                        elif i == rows-3:
                            coefs_arr[i][rows-4] = 2
                            coefs_arr[i][rows-3] = -3
                            coefs_arr[i][rows-2] = 0
                            coefs_arr[i][rows-1] = 1
                            for j in range(rows-4):
                                coefs_arr[i][j] = 0
                            const_vec[i][0] = 0
                        elif i == rows-2:
                            coefs_arr[i][rows-4] = 1
                            coefs_arr[i][rows-3] = 0
                            coefs_arr[i][rows-2] = -3
                            coefs_arr[i][rows-1] = 2
                            for j in range(rows-4):
                                coefs_arr[i][j] = 0
                            const_vec[i][0] = 0
                        elif i == rows-1:
                            coefs_arr[i][rows-4] = 0
                            coefs_arr[i][rows-3] = 1
                            coefs_arr[i][rows-2] = 1
                            coefs_arr[i][rows-1] = 1
                            for j in range(rows-4):
                                coefs_arr[i][j] = 0
                            const_vec[i][0] = self.values[var][dates[-1]]
                    #endregion
                    values_arr = np.matmul(np.linalg.inv(coefs_arr), const_vec)
                    values_lst = list([float(x[0]) for x in values_arr])
                    values_dict = dict(zip(month_dates, values_lst))
                    #region adjusted farvardins
                    if farvardin_adj:
                        for date in dates:
                            year, season = [int(x) for x in date.split('-')]
                            if season == 1:
                                # first adjusted
                                m01 = values_dict[f'{year}-01']*0.6
                                m02 = values_dict[f'{year}-01']*0.4+values_dict[f'{year}-02']*.8
                                m03 = values_dict[f'{year}-03']*1.2
                                total = values_dict[f'{year}-01'] + \
                                    values_dict[f'{year}-02'] + \
                                    values_dict[f'{year}-03']
                                # final adjusted
                                values_dict[f'{year}-01'] = m01 * total / (m01+m02+m03)
                                values_dict[f'{year}-02'] = m02 * total / (m01+m02+m03)
                                values_dict[f'{year}-03'] = m03 * total / (m01+m02+m03)
                    #endregion
                    res.add_data(Data(values={var:values_dict}))
                    if print_progress:
                        prelog = progress(start_time, var_j, vars_no, prelog, var, True, indent+2*subindent)
            else:
                values = {}
                start_year, start_season = [int(x) for x in self.dates[0].split('-')]
                end_year, end_season = [int(x) for x in self.dates[-1].split('-')]
                start_month, end_month = (start_season-1)*3+1, (end_season-1)*3+3
                for year in range(start_year, end_year+1):
                    st = start_month if year == start_year else 1
                    en = end_month+1 if year == end_year else 13
                    for month in range(st,en):
                        date = f'{year}-{month}' if month >9 else f'{year}-0{month}'
                        for var in self.variables():
                            if not var in values.keys(): 
                                values[var] = {}
                            if method=='sum':
                                values[var][date] = self.values[var][f'{year}-{(month-1)//3+1}']/3
                            elif method=='average':
                                values[var][date] = self.values[var][f'{year}-{(month-1)//3+1}']
                    if print_progress:
                        prelog = progress(start_time, year-start_year, end_year-start_year+1,prelog, year, True, indent+2*subindent)
                res.add_data(Data(values=values))
            res = res.to_timeseries(print_progress=print_progress, indent=indent+subindent)
            res.reset_date_type()
            res.complete_dates(print_progress=print_progress, indent=indent+subindent, subindent=subindent)
            return res
        elif self.date_type == 'annual':
            if print_progress:
                start_time, prelog, n = time.perf_counter(), 0, len(self.dates())
            values = {}
            for j, year in enumerate(self.dates):
                for month in range(1,13):
                    date = f'{year}-{month}' if month >9 else f'{year}-0{month}'
                    for var in self.variables():
                        if not var in values.keys(): 
                            values[var] = {}
                        if method=='sum':
                            values[var][date] = self.values[var][year]/12
                        elif method=='average':
                            values[var][date] = self.values[var][year]
                if print_progress:
                    prelog = progress(start_time, j, n, prelog, year, True, indent+2*subindent)
            res = TimeSeries('time', values)
            res.reset_date_type()
            res.complete_dates(print_progress=print_progress, indent=indent+subindent, subindent=subindent)
            return res

    def to_daily(self, print_progress:bool=False, indent:int=0, subindent:int=5)->TimeSeries:
        '''
        convert data to a daily data
        '''
        if print_progress:
            print(' '*indent + 'converting data to daily.')
        if self.date_type == 'daily':
            return self
        elif self.date_type == 'weekly':
            pass    #TODO TimeSeries from weekly to daily
        elif self.date_type == 'monthly':
            y,m = [int(x) for x in self.index()[0].split('-')]
            start = jdatetime.date(y,m, 1)
            y,m = [int(x) for x in self.index()[-1].split('-')]
            days_of_end_month = (jdatetime.date(y,m+1,1)-
                                    jdatetime.date(y,m,1)).days
            end = jdatetime.date(y,m, days_of_end_month)
            new_index, date = [], start
            while (end-date).days >= 0:
                date = date + jdatetime.timedelta(1)
                new_index.append(date)
            values = {}
            for var in self.variables():
                values[var] = {}
            for var in self.variables():
                for i in new_index:
                    index_new = i.strftime('%Y-%m-%d')
                    index_old = f'{i.year}-{i.month}' if i.month>9 else f'{i.year}-0{i.month}'
                    is_in_data = False
                    for j in self.index():
                        if i.year == int(j.split('-')[0]) and \
                            i.month == int(j.split('-')[1]):
                            is_in_data = True
                            break
                    if is_in_data:
                        values[var][index_new] = self.values[var][index_old]
            res = TimeSeries('time', values)
            res.reset_date_type()
            res.complete_dates(print_progress=print_progress, indent=indent+subindent, subindent=subindent)
            return res
        elif self.date_type == 'seasonal':
            pass    #TODO TimeSeries from seasonal to daily
        elif self.date_type == 'annual':
            pass    #TODO TimeSeries from annual to daily

    def to_weekly(self, method:str='average',
                  print_progress:bool=False, indent:int=0, subindent:int=5)->TimeSeries:
        '''
        convert data to a weekly data. This method currently only works on daily time series. 
        '''
        if print_progress:
            print(' '*indent + 'converting data to weekly.')
            start_time, prelog, n = time.perf_counter(), 0, len(self.variables())
        if self.date_type == 'daily':
            values = {}
            for j, var in enumerate(self.variables()):
                values[var] = {}
                w, start = 0, True
                for date in self.dates:
                    sep = '-'
                    y,m,d = [int(x) for x in date.split(sep)]
                    jdate = jdatetime.date(y, m, d)
                    weekday = jdate.weekday()
                    v = self.values[var][date]
                    if weekday == 0:
                        w += 1
                        if not is_nan(v, is_number=True):
                            s, n = v, 1
                        else:
                            s, n = 0, 0
                        start = False
                    elif not start:
                        if not is_nan(v, is_number=True):
                            s += v
                            n += 1
                        if weekday == 6:
                            if n == 0:
                                values[var][date] = math.nan
                            else:
                                values[var][date] = s/n if method=='average' else s
                endweek = (jdate + jdatetime.timedelta(6-weekday)).strftime('%Y-%m-%d')
                if n == 0:
                    values[var][endweek] = math.nan
                else:
                    values[var][endweek] = s/n if method == 'average' else s
                if print_progress:
                    prelog = progress(start_time, j, n, prelog, var, True, indent+subindent)
            return TimeSeries('time', values)

    def to_seasonal(self, method:str='average',
                  print_progress:bool=False, indent:int=0, subindent:int=5)->TimeSeries:
        '''convert data to a seasonal data. This method currently only works on monthly time series.'''
        if print_progress:
            print(' '*indent + 'converting data to seasonal.')
            start_time, prelog, n = time.perf_counter(), 0, len(self.variables())
        if self.date_type == 'monthly':
            values = {var:{} for var in self.variables()}
            for j, var in enumerate(self.variables()):
                sums, numbs, seasons = {}, {}, []
                for date in self.dates:
                    if '-' in date:
                        y, m = date.split('-')
                    elif '/' in date:
                        date.split('/')
                    else:
                        raise ValueError(f'Error! {date} is not a standard date.')
                    seasons.append(season:=f'{y}-{int((int(m)-1)/3)+1}')
                    if not is_nan(x:=self.values[var][date]):
                        if not season in sums:
                            sums[season] = x
                            numbs[season] = 1
                        else:
                            sums[season] += x
                            numbs[season] += 1
                if method == 'sum':
                    values = sums
                elif method == 'average':
                    for season in seasons:
                        try:
                            values[var][season] = sums[season]/numbs[season]
                        except:
                            values[var][season] = math.nan
                else:
                    raise ValueError(f"Error! {method} is not a standard method. standard mathods are 'average' or 'sum'.")
                if print_progress:
                    prelog = progress(start_time, j, n, prelog, var, True, indent+subindent)
            return TimeSeries('time', values)



    def to_annual(self, method:str='average', skip_editing_last_year:bool=True,
                   print_progress:bool=False, indent:int=0, subindent:int=5)->TimeSeries:
        '''
        convert data to a annual data
        '''
        if print_progress:
            print(' '*indent + 'converting data to annual.')
            start_time, prelog, n = time.perf_counter(), 0, len(self.dates)
        values, counts = {}, {}
        for i, date in enumerate(self.dates):
            try:
                year = int(date[:4])
                for var in self.variables():
                    if not var in values.keys():
                        values[var] = {}
                        counts[var] = {}
                    if not is_nan((v:=self.values[var][date])):
                        if year in values[var].keys():
                            values[var][year] += float(v)
                        else:
                            values[var][year] = float(v)
                        if year in counts.keys():
                            counts[var][year] += 1
                        else:
                            counts[var][year] = 1
            except:
                raise ValueError(f'Error! {date} is not standard: yyyy-q or yyyy-mm or yyyy-mm-dd.')
            if print_progress:
                prelog = progress(start_time, i, n, prelog, date, indent+subindent)
        if not skip_editing_last_year:
            if print_progress:
                print(' '*(indent+subindent)+'editing last year')
                start_time, prelog, n= time.perf_counter(), 0, len(values.keys())
            for j, var in enumerate(values.keys()):
                values_1, counts_1 = 0, 0
                for date in self.dates:
                    if date[:4]==str(year) and date.replace(str(year), str(year-1)) in self.dates:
                        if is_nan(self.values[var][date.replace(str(year), str(year-1))], True):
                            values_1 += values[var][year]
                            counts_1 += counts[var][year]
                values[var][year] = values[var][year-1]/values_1 * values[var][year]
                counts[var][year] = counts[var][year-1]/counts_1 * counts[var][year]
                if print_progress:
                    prelog = progress(start_time, j, n, var, True, indent+2*subindent)
        if method == 'average':
            for var in values.keys():
                for year in values[var].keys():
                    values[var][year] /= counts[var][year]
        return TimeSeries(values=values)

    def select_variables(self, variable: str|list[str] = [], print_progress:bool=False, indent:int=0) -> TimeSeries:
        res = super().select_variables(variable, print_progress, indent)
        return TimeSeries('time', res.values)
    
    def select_index(self, index: list, print_progress:bool=False, indent:int=0) -> TimeSeries:
        res = super().select_index(index, print_progress, indent)
        return TimeSeries('time', res.values)

    def add_data(self, new_data: Data = None, print_progress:bool=False, indent:int=0, subindent:int=5) -> TimeSeries:
        super().add_data(new_data, print_progress, indent, subindent)
        self.sort(print_progress=False)
        self.dates = self.index()
        return self

    @staticmethod
    def dates_of_a_period(start_date:str, end_date:str, date_type:str='daily')->list[str]:
        '''
        a function create a list of dates base on type.\n
        '''
        dates = []
        if date_type == 'daily':
            if '-' in start_date and '-' in end_date:
                separator = '-'
                start = jdatetime.date(*[int(p) for p in start_date.split('-')])
                end = jdatetime.date(*[int(p) for p in end_date.split('-')])
            elif '/' in start_date and '/' in end_date:
                separator = '/'
                start = jdatetime.date(*[int(p) for p in start_date.split('/')])
                end = jdatetime.date(*[int(p) for p in end_date.split('/')])
            else:
                raise ValueError(f"Error! seprator must be '-' or '/'.")
            date = start
            while (end-date).days>=0:
                dates.append(date.strftime('%Y-%m-%d').replace('-',separator))
                date += jdatetime.timedelta(1)
        elif date_type == 'weekly':
            if '-' in start_date and '-' in end_date:
                separator = '-'
                start = jdatetime.date(*[int(p) for p in start_date.split('-')])
                end = jdatetime.date(*[int(p) for p in end_date.split('-')])
            elif '/' in start_date and '/' in end_date:
                separator = '/'
                start = jdatetime.date(*[int(p) for p in start_date.split('/')])
                end = jdatetime.date(*[int(p) for p in end_date.split('/')])
            else:
                raise ValueError(f"Error! seprator must be '-' or '/'.")
            date = start
            while (end-date).days>=0:
                dates.append(date.strftime('%Y-%m-%d').replace('-',separator))
                date += jdatetime.timedelta(7)
        elif date_type == 'monthly':
            if '-' in start_date and '-' in end_date:
                separator = '-'
                y_, m_ = [int(p) for p in start_date.split('-')]
                _y, _m = [int(p) for p in end_date.split('-')]
            elif '/' in start_date and '/' in end_date:
                separator = '/'
                y_, m_ = [int(p) for p in start_date.split('/')]
                _y, _m = [int(p) for p in end_date.split('/')]
            else:
                raise ValueError(f"Error! seprator must be '-' or '/'.")
            for y in range(y_, _y+1):
                st = m_ if y==y_ else 1
                en = _m+1 if y==_y else 13
                for m in range(st, en):
                    date = f'{y}{separator}0{m}' if m<10 else f'{y}{separator}{m}'
                    dates.append(date)
        elif date_type == 'seasonal':
            if '-' in start_date and '-' in end_date:
                separator = '-'
                y_, s_ = [int(p) for p in start_date.split('-')]
                _y, _s = [int(p) for p in end_date.split('-')]
            elif '/' in start_date and '/' in end_date:
                separator = '/'
                y_, s_ = [int(p) for p in start_date.split('/')]
                _y, _s = [int(p) for p in end_date.split('/')]
            else:
                raise ValueError(f"Error! seprator must be '-' or '/'.")
            for y in range(y_, _y+1):
                st = s_ if y==y_ else 1
                en = _s+1 if y==_y else 5
                for s in range(st, en):
                    dates.append(f'{y}{separator}{s}')
        elif date_type == 'annual':
            try:
                for date in range(start_date, end_date+1):
                    dates.append(date)
            except:
                raise ValueError(f"Error! {start_date} or {end_date} are not integer.")
        return dates

    def dropna(self, vars: list[str] = []) -> None:
        super().dropna(vars)
        self.reset_date_type()

    fillna_methods = ['last', 'mean', 'growth']
    def fillna(self, variables:str|list[str]=[] , method:str|int|float='last', replace:bool=False):
        '''
        Put a value on place of nans. this value calculate by method: last, mean or growth.
        '''
        if type(variables) == str:
            variable = variables
            last_value, next_value= math.nan, math.nan
            n, index, j = len(self), self.index(), 0
            if not replace:
                self.values[variable + '_'] = {}
            while j < n:
                v = self.values[variable][index[j]]
                try:
                    if is_nan(v):
                        if method == 'last':
                            if not replace:
                                self.values[variable + '_'][index[j]] = last_value
                            else:
                                self.values[variable][index[j]] = last_value
                        elif method == 'mean':
                            k = 1
                            next_value = math.nan
                            while j+k < n:
                                v_ = self.values[variable][index[j+k]]
                                if not is_numeric(v):
                                    raise ValueError(f"Error! value of '{v_}' in variable '{variable}' is not numeric.")
                                try:
                                    if not math.isnan(v_):
                                        next_value = v_
                                        break
                                except:
                                    raise ValueError(f"Error! value of '{v}' in variable '{variable}'.")
                                k+=1
                            k_ = 0
                            if not is_numeric(last_value):
                                raise ValueError(f"Error! value of '{last_value}' in variable '{variable}' is not numeric.")
                            if not is_numeric(next_value):
                                raise ValueError(f"Error! value of '{next_value}' in variable '{variable}' is not numeric.")
                            while k_ < k:
                                if not (math.isnan(last_value) or math.isnan(next_value)):
                                    try:
                                        if not replace:
                                            self.values[variable + '_'][index[j+k_]] = ((k_+1)*last_value+(k-k_)*next_value)/(k+1)
                                        else:
                                            self.values[variable][index[j+k_]] = ((k_+1)*last_value+(k-k_)*next_value)/(k+1)
                                    except:
                                        raise ValueError(f"Error! value of '{v}' in variable '{variable}'.")
                                else:
                                    if not replace:
                                        self.values[variable + '_'][index[j+k_]] = math.nan
                                    else:
                                        self.values[variable][index[j+k_]] = math.nan
                                k_+=1 
                            if not replace:
                                self.values[variable + '_'][index[j+k]] = next_value
                            else:
                                self.values[variable][index[j+k]] = next_value
                            last_value = next_value
                            j += k
                        elif method == 'growth':
                            k = 1
                            next_value = math.nan
                            while j+k < n:
                                v_ = self.values[variable][index[j+k]]
                                if not is_numeric(v):
                                    raise ValueError(f"Error! value of '{v_}' in variable '{variable}' is not numeric.")
                                try:
                                    if not math.isnan(v_):
                                        next_value = v_
                                        break
                                except:
                                    raise ValueError(f"Error! value of '{v}' in variable '{variable}'.")
                                k+=1
                            k_ = 0
                            if not is_numeric(last_value):
                                raise ValueError(f"Error! value of '{last_value}' in variable '{variable}' is not numeric.")
                            if not is_numeric(next_value):
                                raise ValueError(f"Error! value of '{next_value}' in variable '{variable}' is not numeric.")
                            while k_ < k:
                                if not (math.isnan(last_value) or math.isnan(next_value)):
                                    try:
                                        if not replace:
                                            self.values[variable + '_'][index[j+k_]] = last_value * (next_value/last_value)**((k_+1)/(k+1))
                                        else:
                                            self.values[variable][index[j+k_]] = last_value * (next_value/last_value)**((k_+1)/(k+1))
                                    except:
                                        raise ValueError(f"Error! value of '{v}' in variable '{variable}'.")
                                else:
                                    if not replace:
                                        self.values[variable + '_'][index[j+k_]] = math.nan
                                    else:
                                        self.values[variable][index[j+k_]] = math.nan
                                k_+=1 
                            if not replace:
                                self.values[variable + '_'][index[j+k]] = next_value
                            else:
                                self.values[variable][index[j+k]] = next_value
                            last_value = next_value
                            j += k
                        else:
                            if not replace:
                                self.values[variable + '_'][index[j]] = method
                            else:
                                self.values[variable][index[j]] = method
                    else:
                        if not replace:
                            self.values[variable + '_'][index[j]] = v
                        else:
                            self.values[variable][index[j]] = v
                        last_value = v
                    j += 1
                except:
                    raise ValueError(f"Error! value of '{v}' in variable '{variable}', index = {index[j]}.")
        elif type(variables)==list:
            if variables == []:
                variables = self.variables()
            for var in variables:
                self.fillna(var, method, replace)

    #region read data
    @classmethod
    def load(cls, path_file:str)->TimeSeries:
        return super().load(path_file)
        
    @classmethod
    def read_csv(cls, path_file:str, data_type:str='cross', na:any='', index:str='index',
                 variable_number_range:tuple[int]=(),
                 variable_names:list[str] = [], only_names:bool=False, encoding:str='utf8',
                 print_progress:bool=False, indent:int=0, subindent:int=5)->list|TimeSeries:
        data = Data.read_csv(path_file, data_type, na, index, variable_number_range,variable_names,only_names, encoding,print_progress, indent, subindent)
        if only_names:
            return data
        return data.to_timeseries(False)

    @classmethod
    def read_text(cls, path_file:str, data_type:str='cross', na:any='', index:str='index',
                 variable_number_range:tuple[int]=(),
                 variable_names:list[str] = [], only_names:bool=False, seprator:str=',', encoding:str='utf8',print_progress:bool=False,
                 indent:int=0, subindent:int=5)->list|TimeSeries:
        data = Data.read_text(path_file, data_type, na, index, variable_number_range,variable_names,only_names,seprator,encoding,print_progress,indent, subindent)
        if only_names:
            return data
        return data.to_timeseries(False)

    @classmethod
    def read_xls(cls, path_file:str, data_type:str='cross', na:any='', index:str='index',
                 print_progress:bool=False, indent:int=0)->TimeSeries:
        data = Data.read_xls(path_file, data_type, na, index, print_progress, indent)
        data = data.to_timeseries()
        return data

    @classmethod
    def read_excel(cls, path_file:str, sheet:str|int=0, data_type:str='cross', na:any='', index:str='index',
                    first_row:int=0, first_col:int=0, data_only:bool=True, print_progress:bool=False, indent:int=0, subindent:int=5)->TimeSeries:
        data = Data.read_excel(path_file, sheet, data_type, na, index, first_row, first_col, data_only, print_progress, indent, subindent)
        data = data.to_timeseries()
        return data

    @classmethod
    def read_access(cls, path_file:str, table_name:str, variable_names:list[str]=['*'],
                    data_type:str='cross', index:str='index_', na:str='',
                    print_progress:bool=False, indent:int=0, subindnet:int=5)->TimeSeries:
        data = Data.read_access(path_file, table_name, variable_names, data_type, index, na, print_progress, indent, subindnet)
        data = data.to_timeseries()
        return data
    #endregion

    def sort(self, key:str='', ascending:bool=True, print_progress:bool=False, indent:int=0):
        super().sort(key, ascending, print_progress, indent)
        self.dates.sort()

    def add_index(self, index: list, print_progress:bool=False, indent:int=0):
        super().add_index(index, print_progress, indent)
        self.dates = self.index()
        self.complete_dates(print_progress=print_progress, indent=indent*2, subindent=indent)

    def to_growth(self, vars:list[str]=[], lag:int=1, is_average:bool=False, 
                    complete_dates:bool=False, print_progress:bool=False, indent:int=0, subindent:int=5)->TimeSeries:
        '''
        converting data values to growth.
        '''
        if print_progress:
            print(' '*indent+'converting data values to growth')
            start_time, prelog = time.perf_counter(), 0
        if vars==[]:
            vars = self.variables()
        else:
            vars = [var for var in vars if var in self.variables()]
        if complete_dates:
            self.complete_dates(print_progress=print_progress, indent=indent+subindent, subindent=subindent)
        values, v, n = {}, 0, len(vars)
        for j, var in enumerate(vars):
            v += 1
            values[f'gr_{var}'] = {}
            for i in range(len(self.dates)):
                if not is_average:
                    if not (is_nan(self.values[var][self.dates[i-lag]], True) or 
                            is_nan(self.values[var][self.dates[i]], True)):
                        if self.values[var][self.dates[i-lag]] != 0:
                            values[f'gr_{var}'][self.dates[i]] = (self.values[var][self.dates[i]]/
                                                self.values[var][self.dates[i-lag]])-1
                        else:
                            values[f'gr_{var}'][self.dates[i]] = math.nan
                    else:
                        values[f'gr_{var}'][self.dates[i]] = math.nan
                else:
                    if not is_nan(self.values[var][self.dates[i]], True):
                        s, s_lag = 0, 0
                        n, n_lag = 0, 0
                        for l in range(lag):
                            if not(is_nan(self.values[var][self.dates[i-l]],True)):
                                s += self.values[var][self.dates[i-l]]
                                n += 1
                            if not(is_nan(self.values[var][self.dates[i-lag-l]],True)):
                                s_lag += self.values[var][self.dates[i-lag-l]]
                                n_lag += 1
                        if n!=0 and n_lag!=0:
                            values[f'gr_{var}'][self.dates[i]] = (s/n)/(s_lag/n_lag)-1
                        else:
                            values[f'gr_{var}'][self.dates[i]] = math.nan
                    else:
                        values[f'gr_{var}'][self.dates[i]] = math.nan
            if print_progress:
                prelog = progress(start_time, j, n, prelog, var, True, indent+subindent)
        data = TimeSeries(values=values)
        data.complete_dates(print_progress=print_progress, indent=indent+subindent, subindent=subindent)
        return data

    def to_moving_average(self, periods:int, average_period:int=-1, vars:list[str]=[], dates:list[str]=[],
                          print_progress:bool=False, indent:int=0, subindnet:int=5)->TimeSeries:
        '''
        converting data values to moving average
        '''
        if print_progress:
            print(' '*indent+'converting data values to moving average')
            start_time, prelog = time.perf_counter(), 0
        if periods==0:
            return self
        if vars==[]:
            vars = self.variables()
        else:
            vars = [var for var in vars if var in self.variables()]
        if dates ==[]:
            dates = enumerate(self.dates)
        else:
            dates = [(i, date) for i, date in enumerate(self.dates) if date in dates]
        if average_period == -1:
            average_period = periods-1
        elif average_period < 0 or average_period >= periods:
            raise ValueError(f"'{average_period}' must be bettween 0 and {periods-1}.")
        values = {}
        n = len(vars)
        for j, var in enumerate(vars):
            values[var] = {}
            for i, date in dates:
                if i < average_period:
                    values[var][date] = math.nan
                elif i<=len(self.dates) + average_period - periods:
                    vals = [self.values[var][self.dates[i+t-average_period]]
                            for t in range(periods)
                            if not is_nan(self.values[var][self.dates[i+t-average_period]])]
                    if len(vals)>=0:
                        values[var][date] = sum(vals)/len(vals)
                    else:
                        values[var][date] = math.nan
                else:
                    values[var][date] = math.nan
            if print_progress:
                prelog = progress(start_time, j, n, prelog, var, True, indent+subindnet)
        res = TimeSeries(values=values)
        res.dates = self.dates
        res.date_type = self.date_type
        return res

    def to_lead(self, lead:int, variables:list[str]=[],
                print_progress:bool=False, indent:int=0, subindent:int=5)->TimeSeries:
        """
        converting data values to lead.\n
        y_new(t) == y_old(t+lead)
        """
        if print_progress:
            print(' '*indent+f'converting data values to lead {lead}')
        if variables == []:
            variables = self.variables()
        if lead < 0:
            raise ValueError(f"lead ('{lead}') must be positive.")
        new_dates = [Date(self.dates[0], self.date_type).before(l).date for l in range(lead,0,-1)] + self.dates
        values = {var:{} for var in variables}
        for t, date in enumerate(new_dates[:-lead]):
            for var in variables:
                values[var][date] = self.values[var][new_dates[t+lead]]
        for date in new_dates[-lead:]:
            for var in variables:
                values[var][date] = math.nan
        leaded = TimeSeries(values=values)
        leaded.complete_dates(print_progress=print_progress, indent=indent+subindent, subindent=subindent)
        return leaded

    def to_lag(self, lags:list[int], variables:list[str]=[], add_to_data:bool=True,
               print_progress:bool=False, indent:int=0, subindent:int=5)->TimeSeries:
        """
        creating data values to lag.\n
        - y_new(t) = y_old(t-lag) if t>=lag\n
        - y_new(t) = nan if t<lag\n
        """
        if print_progress:
            if len(lags)==1:
                print(' '*indent+f'making data values to lag {lag[0]}')
            else:
                print(' '*indent+f'making data values to lag {lags[0]}...{lags[-1]}')
        if variables == []:
            variables = self.variables()
        for lag in lags:
            if lag < 0:
                raise ValueError(f"lag ('{lag}') must be positive.")
        all_dates = self.dates + (new_dates:=[Date(self.dates[-1], self.date_type).next(l).date for l in range(1, max(lags)+1)])
        lagged_values = {f'l{lag}_{var}':{} for var in variables for lag in lags}
        if add_to_data:
            for var in variables:
                for lag in lags:
                    self.values[f'l{lag}_{var}'] = {}
        # middle of dates (from end to start)
        start, prelog, n = time.perf_counter(), 0, len(lags)
        for l, lag in enumerate(lags):
            for t, date in enumerate(all_dates[:lag-1:-1]):
                for var in variables:
                    if (lagged_date:=all_dates[len(all_dates)-t-lag-1]) in self.dates:
                        lagged_values[f'l{lag}_{var}'][date] = self.values[var][lagged_date]
                    else:
                        lagged_values[f'l{lag}_{var}'][date] = math.nan
                    if add_to_data:
                        self.values[f'l{lag}_{var}'][date] = lagged_values[f'l{lag}_{var}'][date]
            if print_progress:
                prelog = progress(start, l, n, prelog, 'lag', False, indent+subindent)
        # start of dates (values are nan)
        for lag in lags:
            for date in self.dates[:lag]:
                for var in variables:
                    lagged_values[f'l{lag}_{var}'][date] = math.nan
                    if add_to_data:
                        self.values[f'l{lag}_{var}'][date] = math.nan
        if add_to_data:
            for var in self.variables():
                for date in new_dates:
                    if not (var[0] == 'l' and
                                 var.split('_')[0][1:].isdigit() and
                                 '_'.join(var.split('_')[1:]) in variables):
                        self.values[var][date] = math.nan
            self.dates = self.index()
        lagged = TimeSeries(values=lagged_values)
        lagged.complete_dates(print_progress=print_progress, indent=indent+subindent, subindent=subindent)
        return lagged

    def drop_index(self, index:list|any, print_progress:bool=False, indent:int=0, subindent:int=5):
        super().drop_index(index, print_progress, indent, subindent)
        if not isinstance(index, list):
            index = [index]
        for date in index:
            self.dates.remove(date)

class DataBase:
    data_types = ['cross', 'time']
    date_types = ['daily', 'weekly', 'monthly', 'seasonal', 'annual']
    def __init__(self, tables:dict[str:Data|TimeSeries]) -> None:
        self.tables = tables

    def dump(self, path:str):
        with open(path, mode='wb') as f:
            pickle.dump(self, f)

    @classmethod
    def load(cls, path:str)->DataBase:
        with open(path, mode='rb') as f:
            return pickle.load(f)
        
    def table_list(self)->list[str]:
        return list(self.tables.keys())
    
    def variable_list(self)->list[tuple[str, Data]]:
        return [(var,table) for table in self.table_list() for var in self.tables[table].variables()]
    
    def query(self, variables:list[tuple[str,Data]])->Data:
        values = {var:{} for var, _ in variables}
        is_timeseries = min([isinstance(table, TimeSeries) for _, table in variables])
        if is_timeseries:
            type_tables = list({table.date_type for _, table in variables})
        else:
            type_tables = list({table.type for _, table in variables})
        if len(type_tables)>1:
            raise ValueError(f"tables must be a type, but tables are types of {str(type_tables)[1:-1]}")
        index = [i for _, table in variables for i in table.index()]
        for var, table in variables():
            for i in index:
                if i in table.index():
                    values[var][i] = table.values[var][i]
                else:
                    values[var][i] = math.nan
        if is_timeseries:
            return TimeSeries(values=values)
        else:
            return Data(values= values)